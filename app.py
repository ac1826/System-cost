import io
import re
import unicodedata
import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title='TSC 计算模型', layout='wide')

st.title('计算模型')

HIDE_COLS = ['半成品入库量', '调整后实际量', '调整后实际额', '碎肉量', '辅助']
FMT_DISPLAY = {
    '修形前原料综合耗用单价': '{:.2f}',
    '半成品原料成本': '{:.2f}',
    '半成品修形人工成本': '{:.2f}',
    '半成品总成本': '{:.2f}',
    '修形利用率': '{:.0%}',
    '损耗率': '{:.0%}',
}

with st.sidebar:
    st.header('输入文件')
    file_compare = st.file_uploader('基础表', type=['xlsx'])
    file_rawlist = st.file_uploader('原料清单.xlsx', type=['xlsx'])
    file_q3 = st.file_uploader('系统成本', type=['xlsx'])
    file_map = st.file_uploader('半成品对应成品.xlsx', type=['xlsx'])

def _find_header_row(df, keyword):
    for i in range(min(len(df), 50)):
        row = df.iloc[i].astype(str).apply(_clean_colname)
        if row.str.contains(_clean_colname(keyword)).any():
            return i
    return None


def _normalize_mat(val):
    if pd.isna(val):
        return ''
    if isinstance(val, (int,)):
        return str(val)
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val).rstrip('0').rstrip('.')
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s


def _clean_colname(val):
    s = str(val).strip()
    s = re.sub(r'\s+', '', s)
    return s


def _load_rawlist(rawlist_file):
    sheets = pd.read_excel(rawlist_file, sheet_name=None)
    frames = []
    for df in sheets.values():
        df = df.copy()
        df.columns = [_clean_colname(c) for c in df.columns]
        if '原料号' in df.columns and '部位' in df.columns:
            frames.append(df[['原料号', '部位']])
    if not frames:
        return pd.DataFrame(columns=['原料号', '部位'])
    raw = pd.concat(frames, ignore_index=True)
    raw['原料号'] = raw['原料号'].apply(_normalize_mat)
    raw['部位'] = raw['部位'].fillna('').astype(str).str.strip()
    return raw


def _category_from_part(part):
    if part == '腿类':
        return '腿肉'
    if part == '胸类':
        return '胸肉'
    return '其他'


def _ensure_columns(df, cols):
    cols_set = set(df.columns)
    missing = [c for c in cols if c not in cols_set]
    if missing:
        raise ValueError(f'缺少列: {missing}')


def _normalize_columns(df):
    col_map = {}
    for col in df.columns:
        key = _clean_colname(col)
        if key not in col_map:
            col_map[col] = key
    df = df.rename(columns=col_map)
    # handle duplicate names by keeping the first occurrence
    df = df.loc[:, ~df.columns.duplicated()]
    return df


def _ensure_and_map_columns(df, required):
    # map cleaned names to existing columns
    clean_map = {}
    for col in df.columns:
        clean = _clean_colname(col)
        if clean not in clean_map:
            clean_map[clean] = col

    def _find_col_by_keywords(keywords):
        for col in df.columns:
            clean = _clean_colname(col)
            for kw in keywords:
                if kw in clean:
                    return col
        return None

    alias_map = {
        '物料号': ['物料编号', '物料编码', '物料代码', '物料'],
        '物料描述': ['物料名称', '品名', '物料规格'],
        '原料号': ['原料编号', '原料编码', '原料代码', '原料'],
        '原料描述': ['原料名称', '原料规格', '原料品名'],
        '入库数量': ['入库量', '入库数量(kg)', '入库数量（kg）'],
        '入库金额': ['入库金额(元)', '入库金额（元）', '入库金额含税'],
        '实际数量': ['实际量', '实际数量(kg)', '实际数量（kg）'],
        '实际金额': ['实际金额(元)', '实际金额（元）', '实际金额含税'],
        '配方数量': ['配方量', '配方用量', '配方数量(kg)', '配方数量（kg）'],
    }

    # rename columns to required names where possible
    rename_map = {}
    for req in required:
        if req in df.columns:
            continue
        clean_req = _clean_colname(req)
        if clean_req in clean_map:
            rename_map[clean_map[clean_req]] = req
            continue
        aliases = alias_map.get(req, [])
        alias_clean = [_clean_colname(a) for a in aliases]
        col = _find_col_by_keywords(alias_clean + [clean_req])
        if col:
            rename_map[col] = req
    if rename_map:
        df = df.rename(columns=rename_map)

    # fallback: any column containing "物料号"
    if '物料号' not in df.columns:
        col = _find_col_by_keywords(['物料号', '物料'])
        if col:
            df = df.rename(columns={col: '物料号'})

    _ensure_columns(df, required)
    return df


def _to_num(s):
    return pd.to_numeric(s, errors='coerce').fillna(0.0)


def _find_tsc_value(tsc_file, sheet_name, material_no, row_label):
    df = pd.read_excel(_as_bytes_io(tsc_file), sheet_name=sheet_name, header=None)
    # Find column index for 综合单价
    header_row = _find_header_row(df, '综合单价')
    if header_row is None:
        return None
    header = [_clean_colname(c) for c in df.iloc[header_row].astype(str).tolist()]
    try:
        price_col = header.index('综合单价')
    except ValueError:
        return None
    # Locate row by material no (col 3) and label (col 5)
    mat_col = 2
    label_col = 4
    for i in range(header_row + 1, len(df)):
        mat = _normalize_mat(df.iat[i, mat_col])
        label = str(df.iat[i, label_col]).strip()
        if mat == material_no and label == row_label:
            val = df.iat[i, price_col]
            try:
                return float(val)
            except Exception:
                return None
    return None


def _find_tsc_metrics(tsc_file, sheet_name, material_no, row_label):
    df = pd.read_excel(_as_bytes_io(tsc_file), sheet_name=sheet_name, header=None)
    header_row = _find_header_row(df, '修形前原料综合耗用单价')
    if header_row is None:
        return None
    header = [_clean_colname(c) for c in df.iloc[header_row].astype(str).tolist()]
    header_idx = {}
    for i, h in enumerate(header):
        if h not in header_idx:
            header_idx[h] = i
    alias_map = {
        '修形前原料综合耗用单价': ['修形前原料综合耗用单价'],
        '修形利用率': ['修形利用率'],
        '损耗率': ['损耗率'],
        '半成品原料成本': ['半成品原料成本'],
        '半成品修形人工成本': ['半成品修形人工成本', '人工', '人工成本'],
        '半成品总成本': ['半成品总成本'],
    }
    cols = {}
    for key, aliases in alias_map.items():
        cols[key] = None
        for a in aliases:
            if a in header_idx:
                cols[key] = header_idx[a]
                break
    mat_col = 2
    label_col = 4
    for i in range(header_row + 1, len(df)):
        mat = _normalize_mat(df.iat[i, mat_col])
        label = str(df.iat[i, label_col]).strip()
        if mat == material_no and label == row_label:
            out = {}
            for k, c in cols.items():
                if c is None:
                    out[k] = None
                    continue
                val = df.iat[i, c]
                try:
                    out[k] = float(val)
                except Exception:
                    out[k] = None
            return out
    return None


def _get_tsc_raw_columns(tsc_file, sheet_name):
    df = pd.read_excel(_as_bytes_io(tsc_file), sheet_name=sheet_name, header=None)
    header_row = _find_header_row(df, '综合单价')
    if header_row is None or header_row == 0:
        return [], None, None, {}
    raw_row = header_row - 1
    raw_cols = []
    spec_map = {}
    for c in range(df.shape[1]):
        val = df.iat[raw_row, c]
        if pd.isna(val):
            continue
        s = str(val).strip()
        if s.isdigit():
            raw_cols.append((c, s))
            spec_val = df.iat[header_row, c]
            spec_map[s] = str(spec_val).strip() if not pd.isna(spec_val) else ''
    comp_col = None
    for c in range(df.shape[1]):
        v = df.iat[header_row, c]
        if str(v).strip() == '综合单价':
            comp_col = c
            break
    return raw_cols, comp_col, df, spec_map


def _find_tsc_row_values(df, material_no, row_label, raw_cols, comp_col):
    if df is None:
        return {}
    mat_col = 2
    label_col = 4
    for i in range(len(df)):
        mat = _normalize_mat(df.iat[i, mat_col])
        label = str(df.iat[i, label_col]).strip()
        if mat == material_no and label == row_label:
            out = {}
            for c, code in raw_cols:
                try:
                    out[code] = float(df.iat[i, c])
                except Exception:
                    out[code] = None
            if comp_col is not None:
                try:
                    out['综合单价'] = float(df.iat[i, comp_col])
                except Exception:
                    out['综合单价'] = None
            return out
    return {}


def _load_compare_df(compare_file, sheet_name=None):
    df = pd.read_excel(compare_file, sheet_name=sheet_name or 0)
    df = _normalize_columns(df)
    if '物料号' in df.columns:
        return df
    raw = pd.read_excel(compare_file, sheet_name=sheet_name or 0, header=None)
    header_row = _find_header_row(raw, '物料号')
    if header_row is None:
        header_row = 0
    header = [_clean_colname(c) for c in raw.iloc[header_row].astype(str).tolist()]
    data = raw.iloc[header_row + 1 :].copy()
    data.columns = header
    data = _normalize_columns(data)
    if '物料号' not in data.columns and data.shape[1] > 0:
        # last resort: treat first column as 物料号
        data = data.rename(columns={data.columns[0]: '物料号'})
    return data


def _drop_hidden_cols(df):
    df = df.drop(columns=[c for c in HIDE_COLS if c in df.columns])
    df = df.drop(columns=['行序'], errors='ignore')
    return df


def _month_label_from_filename(name):
    base = name.rsplit('.', 1)[0]
    if '_' not in base:
        return '11月'
    suffix = base.split('_', 1)[1]
    # 优先识别 YYYYMM / YYMM 口径（如 2601 -> 1月）
    m4 = re.search(r'(\d{4})', suffix)
    if m4:
        mm = int(m4.group(1)[-2:])
        if 1 <= mm <= 12:
            return f'{mm}月'
    # 兜底：直接取 1-2 位月份
    m = re.search(r'(\d{1,2})', suffix)
    if not m:
        return '11月'
    mm = int(m.group(1))
    if 1 <= mm <= 12:
        return f'{mm}月'
    return '11月'


def _quarter_label_from_filename(name):
    base = name.rsplit('.', 1)[0]
    m = re.search(r'(Q[1-4])', base, re.IGNORECASE)
    if not m:
        return 'Q3'
    return m.group(1).upper()


def _resolve_tsc_quarter_label(tsc_file, preferred='Q3'):
    preferred_q = str(preferred or '').strip().upper()
    if preferred_q not in {'Q1', 'Q2', 'Q3', 'Q4'}:
        preferred_q = 'Q3'

    scores = {q: 0 for q in ['Q1', 'Q2', 'Q3', 'Q4']}
    pat = re.compile(r'^(Q[1-4])(实际单价|规格占比)$', re.IGNORECASE)

    for sn in ['腿肉TSC', '胸肉TSC']:
        try:
            df = pd.read_excel(_as_bytes_io(tsc_file), sheet_name=sn, header=None)
        except Exception:
            continue
        if df is None or df.empty or df.shape[1] <= 4:
            continue
        labels = df.iloc[:, 4].dropna().astype(str).tolist()
        for lab in labels:
            m = pat.match(_clean_colname(lab).upper())
            if m:
                scores[m.group(1).upper()] += 1

    if hasattr(tsc_file, 'seek'):
        try:
            tsc_file.seek(0)
        except Exception:
            pass

    if scores.get(preferred_q, 0) > 0:
        return preferred_q
    best_q = max(scores, key=scores.get)
    if scores.get(best_q, 0) > 0:
        return best_q
    return preferred_q


def _plant_name_from_rawlist(name):
    base = name.rsplit('.', 1)[0]
    if '_' not in base:
        return base
    return base.split('_', 1)[1]


def _month_code_from_filename(name):
    base = name.rsplit('.', 1)[0]
    m = re.search(r'(\d{4})', base)
    if m:
        return m.group(1)
    m = re.search(r'(\d{1,2})', base)
    if m:
        return str(int(m.group(1)))
    return ''

def _as_bytes_io(file_obj):
    if hasattr(file_obj, 'getvalue'):
        return io.BytesIO(file_obj.getvalue())
    if hasattr(file_obj, 'seek'):
        file_obj.seek(0)
    return file_obj


def _load_mapping(map_file, sheet_name=None):
    map_source = _as_bytes_io(map_file)
    if hasattr(map_file, 'seek'):
        map_file.seek(0)
    def _extract(df):
        header_row = None
        cat_col = None
        semi_col = None
        prod_col = None
        for i in range(min(len(df), 20)):
            row = df.iloc[i].astype(str).fillna('').apply(_clean_colname)
            if row.str.contains('半成品').any() and row.str.contains('对应成品').any():
                header_row = i
                semi_col = row[row.str.contains('半成品')].index[0]
                prod_col = row[row.str.contains('对应成品')].index[0]
                if row.str.contains('分类').any():
                    cat_col = row[row.str.contains('分类')].index[0]
                else:
                    cat_col = semi_col - 1 if semi_col is not None and semi_col - 1 >= 0 else None
                break
        if header_row is None:
            return []
        mappings = []
        current_cat = None
        current_semi = None
        current_semi_desc = None
        start = header_row + 1
        for _, row in df.iloc[start:].iterrows():
            cat = row.iloc[cat_col] if cat_col is not None else None
            semi = row.iloc[semi_col] if semi_col is not None else None
            semi_desc = row.iloc[semi_col + 1] if semi_col is not None and semi_col + 1 < len(row) else None
            prod = row.iloc[prod_col] if prod_col is not None else None
            prod_desc = row.iloc[prod_col + 1] if prod_col is not None and prod_col + 1 < len(row) else None
            if pd.notna(cat) and str(cat).strip() != '':
                current_cat = str(cat).strip()
            if pd.notna(semi) and str(semi).strip() != '':
                current_semi = _normalize_mat(semi)
                current_semi_desc = str(semi_desc).strip() if pd.notna(semi_desc) else ''
            if pd.notna(prod) and str(prod).strip() != '':
                mappings.append(
                    (
                        current_cat,
                        current_semi,
                        current_semi_desc,
                        _normalize_mat(prod),
                        str(prod_desc).strip() if pd.notna(prod_desc) else '',
                    )
                )
        return mappings

    def _extract_compact(df):
        # compact plant-sheet layout (no explicit header):
        # col0 分类(可空，向下填充) | col1 半成品 | col2 半成品描述 | col3 对应成品 | col4 成品描述
        mappings = []
        current_cat = None
        for _, row in df.iterrows():
            if len(row) < 4:
                continue
            cat = row.iloc[0] if len(row) > 0 else None
            semi = row.iloc[1] if len(row) > 1 else None
            semi_desc = row.iloc[2] if len(row) > 2 else None
            prod = row.iloc[3] if len(row) > 3 else None
            prod_desc = row.iloc[4] if len(row) > 4 else None
            if pd.notna(cat) and str(cat).strip() != '':
                current_cat = str(cat).strip()
            if pd.isna(semi) or str(semi).strip() == '':
                continue
            if pd.isna(prod) or str(prod).strip() == '':
                continue
            mappings.append(
                (
                    current_cat,
                    _normalize_mat(semi),
                    str(semi_desc).strip() if pd.notna(semi_desc) else '',
                    _normalize_mat(prod),
                    str(prod_desc).strip() if pd.notna(prod_desc) else '',
                )
            )
        return mappings

    # try specified sheet first (if provided)
    if sheet_name:
        try:
            df = pd.read_excel(map_source, sheet_name=sheet_name, header=None)
            mappings = _extract(df)
            if mappings:
                return mappings
            mappings = _extract_compact(df)
            if mappings:
                return mappings
        except Exception:
            pass

    # scan all sheets for headers and merge
    try:
        map_source = _as_bytes_io(map_file)
        sheets = pd.read_excel(map_source, sheet_name=None, header=None)
    except Exception:
        return []
    merged = []
    for df in sheets.values():
        mappings = _extract(df)
        if mappings:
            merged.extend(mappings)
            continue
        mappings = _extract_compact(df)
        if mappings:
            merged.extend(mappings)
    if merged:
        return merged
    return []


def _load_semi_category(map_file, sheet_name=None):
    # prefer openpyxl to handle merged cells reliably
    try:
        wb = openpyxl.load_workbook(_as_bytes_io(map_file), data_only=True)
        # 扫描全部工作表，并优先使用“汇总”页口径（半成品->分类）
        all_names = list(wb.sheetnames)
        ordered_names = []
        if '汇总' in all_names:
            ordered_names.append('汇总')
        if sheet_name and sheet_name in all_names and sheet_name not in ordered_names:
            ordered_names.append(sheet_name)
        for n in all_names:
            if n not in ordered_names:
                ordered_names.append(n)

        merged = {}
        for name in ordered_names:
            ws = wb[name]
            header_row = None
            semi_col = None
            cat_col = None
            for r in range(1, 21):
                for c in range(1, 31):
                    val = ws.cell(r, c).value
                    if val is None:
                        continue
                    s = _clean_colname(str(val))
                    if '半成品' in s:
                        semi_col = c
                    if '分类' in s:
                        cat_col = c
                if semi_col and cat_col:
                    header_row = r
                    break
                semi_col = None
                cat_col = None
            if not header_row:
                # compact plant-sheet layout fallback:
                # col0 分类(可空向下填充), col1 半成品
                current_cat = None
                for r in range(1, ws.max_row + 1):
                    cat = ws.cell(r, 1).value
                    semi = ws.cell(r, 2).value
                    if cat is not None and str(cat).strip() != '':
                        current_cat = str(cat).strip()
                    if semi is not None and str(semi).strip() != '' and current_cat:
                        semi_key = _normalize_mat(semi)
                        if str(semi_key).startswith('3900'):
                            # 汇总页优先；其余页仅在未命中时补充
                            if name == '汇总' or semi_key not in merged:
                                merged[semi_key] = current_cat
                continue
            current_cat = None
            for r in range(header_row + 1, ws.max_row + 1):
                cat = ws.cell(r, cat_col).value
                semi = ws.cell(r, semi_col).value
                if cat is not None and str(cat).strip() != '':
                    current_cat = str(cat).strip()
                if semi is not None and str(semi).strip() != '' and current_cat:
                    semi_key = _normalize_mat(semi)
                    if str(semi_key).startswith('3900'):
                        if name == '汇总' or semi_key not in merged:
                            merged[semi_key] = current_cat
        if merged:
            return merged
    except Exception:
        pass

    return {}


def _load_finished_pairs_from_system(cost_file):
    # supplement mapping pairs from uploaded system-cost workbook (e.g., 成品-Q4)
    try:
        sheets = pd.read_excel(_as_bytes_io(cost_file), sheet_name=None)
    except Exception:
        return []
    pairs = []
    for sname, df in sheets.items():
        if '成品' not in str(sname):
            continue
        if df.shape[1] < 4:
            continue
        prod_col = df.columns[0]
        prod_desc_col = df.columns[1] if df.shape[1] > 1 else None
        semi_col = df.columns[3]
        semi_desc_col = df.columns[4] if df.shape[1] > 4 else None
        for _, row in df.iterrows():
            prod = _normalize_mat(row.get(prod_col))
            semi = _normalize_mat(row.get(semi_col))
            if not prod or not semi:
                continue
            if not str(prod).strip() or not str(semi).startswith('3900'):
                continue
            if not re.match(r'^[A-Za-z]', str(prod)):
                continue
            pairs.append(
                (
                    None,
                    semi,
                    str(row.get(semi_desc_col)).strip() if semi_desc_col is not None and pd.notna(row.get(semi_desc_col)) else '',
                    prod,
                    str(row.get(prod_desc_col)).strip() if prod_desc_col is not None and pd.notna(row.get(prod_desc_col)) else '',
                )
            )
    # deduplicate by (semi, prod)
    dedup = {}
    for rec in pairs:
        dedup[(rec[1], rec[3])] = rec
    return list(dedup.values())


def _build_output_filename(compare_name, rawlist_name):
    plant = _plant_name_from_rawlist(rawlist_name)
    month_code = _month_code_from_filename(compare_name)
    if month_code:
        return f'{plant}系统成本-{month_code}.xlsx'
    month_label = _month_label_from_filename(compare_name)
    return f'{plant}{month_label}系统成本.xlsx'


def compute(compare_file, rawlist_file, q3_file, map_file, sheet_name=None, month_label='11月', quarter_label='Q3', month_code=''):
    df = _load_compare_df(compare_file, sheet_name=sheet_name)
    if '物料号' not in df.columns and df.shape[1] > 0:
        candidates = [c for c in df.columns if '物料' in str(c)]
        if candidates:
            df = df.rename(columns={candidates[0]: '物料号'})
        else:
            df = df.rename(columns={df.columns[0]: '物料号'})

    required = ['物料号', '物料描述', '原料号', '原料描述', '入库数量', '入库金额', '实际数量', '实际金额', '配方数量']
    df = _ensure_and_map_columns(df, required)
    if '物料号' not in df.columns and df.shape[1] > 0:
        df['物料号'] = df.iloc[:, 0]

    rawlist = _load_rawlist(rawlist_file)
    raw_map = dict(zip(rawlist['原料号'], rawlist['部位']))
    raw_set = set(raw_map.keys())

    df['物料号'] = df['物料号'].apply(_normalize_mat)

    df['原料号'] = df['原料号'].apply(_normalize_mat)
    df['原料描述'] = df['原料描述'].fillna('').astype(str).str.strip()
    df['物料描述'] = df['物料描述'].fillna('').astype(str).str.strip()

    df['入库数量'] = _to_num(df['入库数量'])
    df['入库金额'] = _to_num(df['入库金额'])
    df['实际数量'] = _to_num(df['实际数量'])
    df['实际金额'] = _to_num(df['实际金额'])
    df['配方数量'] = _to_num(df['配方数量'])
    # Fixed business rule: always read quarter metrics as Q3.
    quarter_label = 'Q3'

    # 入库数量从物料号主行提取（原料号为空）
    header_in_qty = (
        df[df['原料号'].isna() | (df['原料号'] == '')]
        .groupby('物料号', as_index=False)['入库数量']
        .max()
        .rename(columns={'入库数量': '入库数量_主行'})
    )

    # 成品表：按“成品物料号 + 半成品原料号”明细口径生成（数据来源仅第一个文件）
    finished_df = pd.DataFrame(
        columns=['物料号', '物料描述', '入库数量', '原料号', '原料描述', '实际数量', '实际金额', '配方数量']
    )
    plant_name = _plant_name_from_rawlist(rawlist_file.name)
    mappings = _load_mapping(map_file, sheet_name=plant_name)
    # supplement missing mapping pairs from uploaded system-cost file
    extra_pairs = _load_finished_pairs_from_system(q3_file)
    if extra_pairs:
        existing = {(m[1], m[3]) for m in mappings}
        for rec in extra_pairs:
            if (rec[1], rec[3]) not in existing:
                mappings.append(rec)
                existing.add((rec[1], rec[3]))
    semi_category = _load_semi_category(map_file, sheet_name=plant_name)
    if mappings:
        map_rows = []
        for cat, semi, semi_desc, prod, prod_desc in mappings:
            if not semi or not prod:
                continue
            map_rows.append(
                {
                    '物料号': prod,
                    '物料描述': prod_desc,
                    '原料号': semi,
                    '原料描述': semi_desc,
                }
            )
            # semi_category 专用于“半成品->分类”映射，来自 _load_semi_category（对应表“半成品/分类”列）
            # 这里不再用 _load_mapping 的 cat 覆盖，避免把口径带偏。
        map_df = pd.DataFrame(map_rows)
        map_df['物料号'] = map_df['物料号'].apply(_normalize_mat)
        map_df['原料号'] = map_df['原料号'].apply(_normalize_mat)

        df_work = df.copy()
        df_work['物料号_norm'] = df_work['物料号'].apply(_normalize_mat)
        df_work['原料号_norm'] = df_work['原料号'].apply(_normalize_mat)

        # 成品主行入库数量（原料号为空）
        in_by_prod = (
            df_work[
                df_work['原料号'].isna()
                | (df_work['原料号'].astype(str).str.strip() == '')
            ]
            .groupby('物料号_norm', as_index=False)['入库数量']
            .sum()
            .rename(columns={'物料号_norm': '物料号', '入库数量': '入库数量'})
        )

        # 成品明细（原料号3900开头）
        pair_sums = (
            df_work[
                df_work['原料号_norm'].astype(str).str.startswith('3900')
                & df_work['物料号_norm'].astype(str).str.match(r'^[A-Za-z]')
            ]
            .groupby(['物料号_norm', '原料号_norm'], as_index=False)[['实际数量', '实际金额', '配方数量']]
            .sum()
            .rename(columns={'物料号_norm': '物料号', '原料号_norm': '原料号'})
        )
        pair_sums['入库数量_回退'] = 0.0
        source_mode = '成品明细'

        # 若第一个文件缺少成品编码，回退到“按半成品汇总后映射”
        if pair_sums.empty:
            source_mode = '半成品回推'
            semi_base = df_work[df_work['物料号_norm'].astype(str).str.startswith('3900')].copy()
            semi_in = (
                semi_base[
                    semi_base['原料号'].isna()
                    | (semi_base['原料号'].astype(str).str.strip() == '')
                ]
                .groupby('物料号_norm', as_index=False)['入库数量']
                .sum()
                .rename(columns={'物料号_norm': '原料号', '入库数量': '入库数量_回退'})
            )
            semi_detail = (
                semi_base[
                    semi_base['原料号'].notna()
                    & (semi_base['原料号'].astype(str).str.strip() != '')
                    & (semi_base['原料号'].astype(str).str.strip() != '人工费用')
                ]
                .groupby('物料号_norm', as_index=False)[['实际数量', '实际金额', '配方数量']]
                .sum()
                .rename(columns={'物料号_norm': '原料号'})
            )
            semi_totals = semi_in.merge(semi_detail, on='原料号', how='outer')
            pair_sums = map_df[['物料号', '原料号']].merge(semi_totals, on='原料号', how='left')

        # Build base pairs:
        # - 成品明细口径: trust compare-file product+3900 pairs directly
        # - 半成品回推口径: use mapped pairs (already expanded above)
        if source_mode == '成品明细':
            finished_df = pair_sums.copy()
            # append mapping-only pairs to keep compatibility, values will be filled below (mostly 0)
            if not map_df.empty:
                map_keys = map_df[['物料号', '原料号']].drop_duplicates()
                cur_keys = finished_df[['物料号', '原料号']].drop_duplicates()
                extra_keys = map_keys.merge(cur_keys, on=['物料号', '原料号'], how='left', indicator=True)
                extra_keys = extra_keys[extra_keys['_merge'] == 'left_only'][['物料号', '原料号']]
                if not extra_keys.empty:
                    extra_keys['实际数量'] = 0.0
                    extra_keys['实际金额'] = 0.0
                    extra_keys['配方数量'] = 0.0
                    extra_keys['入库数量_回退'] = 0.0
                    finished_df = pd.concat([finished_df, extra_keys], ignore_index=True)
        else:
            finished_df = map_df.merge(pair_sums, on=['物料号', '原料号'], how='left')

        # Fill product/raw descriptions from mapping first
        if not map_df.empty:
            finished_df = finished_df.merge(
                map_df[['物料号', '原料号', '物料描述', '原料描述']].drop_duplicates(['物料号', '原料号']),
                on=['物料号', '原料号'],
                how='left',
            )
        else:
            finished_df['物料描述'] = ''
            finished_df['原料描述'] = ''

        # Product description fallback from compare file header rows
        prod_desc_map = (
            df_work[
                (df_work['物料号_norm'].astype(str).str.match(r'^[A-Za-z]'))
                & df_work['物料描述'].notna()
                & (df_work['物料描述'].astype(str).str.strip() != '')
            ][['物料号_norm', '物料描述']]
            .drop_duplicates('物料号_norm')
            .rename(columns={'物料号_norm': '物料号', '物料描述': '物料描述_主行'})
        )
        finished_df = finished_df.merge(prod_desc_map, on='物料号', how='left')
        finished_df['物料描述'] = finished_df.apply(
            lambda r: (
                r['物料描述']
                if (pd.notna(r.get('物料描述')) and str(r.get('物料描述')).strip() != '')
                else r.get('物料描述_主行', '')
            ),
            axis=1,
        )
        finished_df = finished_df.drop(columns=['物料描述_主行'], errors='ignore')

        finished_df = finished_df.merge(
            in_by_prod.rename(columns={'入库数量': '入库数量_成品'}),
            on='物料号',
            how='left',
        )
        finished_df['入库数量'] = finished_df.get('入库数量_成品', 0).fillna(0)
        if '入库数量_回退' in finished_df.columns:
            finished_df['入库数量'] = finished_df['入库数量'].where(
                finished_df['入库数量'] != 0,
                finished_df['入库数量_回退'].fillna(0),
            )
        finished_df = finished_df.drop(columns=['入库数量_成品', '入库数量_回退'], errors='ignore')
        for c in ['入库数量', '实际数量', '实际金额', '配方数量']:
            if c in finished_df.columns:
                finished_df[c] = finished_df[c].fillna(0)

        # keep only produced/used rows to match system-cost finished-sheet granularity
        finished_df = finished_df[
            (finished_df['入库数量'] != 0)
            | (finished_df['实际数量'] != 0)
            | (finished_df['实际金额'] != 0)
            | (finished_df['配方数量'] != 0)
        ].copy()

        # 补齐原料描述（优先映射表，其次第一个文件半成品主行）
        semi_desc = (
            df_work[
                (df_work['原料号'].isna() | (df_work['原料号'].astype(str).str.strip() == ''))
                & df_work['物料号_norm'].astype(str).str.startswith('3900')
            ][['物料号_norm', '物料描述']]
            .dropna(subset=['物料号_norm'])
            .drop_duplicates('物料号_norm')
            .rename(columns={'物料号_norm': '原料号', '物料描述': '原料描述_半成品'})
        )
        finished_df = finished_df.merge(semi_desc, on='原料号', how='left')
        finished_df['原料描述'] = finished_df.apply(
            lambda r: (
                r['原料描述']
                if (pd.notna(r.get('原料描述')) and str(r.get('原料描述')).strip() != '')
                else r.get('原料描述_半成品', '')
            ),
            axis=1,
        )
        finished_df = finished_df.drop(columns=['原料描述_半成品'], errors='ignore')

        # 扩展列：系列 / 实际收得率 / 配方收得率 / 部位 / 月份
        finished_df['系列'] = finished_df['原料号'].apply(lambda x: semi_category.get(_normalize_mat(x), ''))

        def _part_from_raw(raw):
            p = raw_map.get(_normalize_mat(raw), '')
            if p == '胸类':
                return '胸肉'
            if p == '腿类':
                return '腿肉'
            return ''

        finished_df['部位'] = finished_df['原料号'].apply(_part_from_raw)
        if month_code and str(month_code).strip() != '':
            finished_df['月份'] = str(month_code).strip()
        else:
            m_month = re.search(r'(\d+)', str(month_label))
            finished_df['月份'] = m_month.group(1) if m_month else str(month_label).replace('月', '')
        finished_df['实际收得率'] = finished_df.apply(
            lambda r: (r['实际数量'] / r['入库数量']) if r['入库数量'] else 0,
            axis=1,
        )
        finished_df['配方收得率'] = finished_df.apply(
            lambda r: (r['配方数量'] / r['入库数量']) if r['入库数量'] else 0,
            axis=1,
        )
        finished_df['来源口径'] = source_mode

        # 删除“关键指标全为空/0”的行
        s_actual_qty = pd.to_numeric(finished_df.get('实际数量', 0), errors='coerce').fillna(0)
        s_actual_amt = pd.to_numeric(finished_df.get('实际金额', 0), errors='coerce').fillna(0)
        s_formula_qty = pd.to_numeric(finished_df.get('配方数量', 0), errors='coerce').fillna(0)
        s_actual_yield = pd.to_numeric(finished_df.get('实际收得率', 0), errors='coerce').fillna(0)
        s_formula_yield = pd.to_numeric(finished_df.get('配方收得率', 0), errors='coerce').fillna(0)
        all_empty_mask = (
            (s_actual_qty == 0)
            & (s_actual_amt == 0)
            & (s_formula_qty == 0)
            & (s_actual_yield == 0)
            & (s_formula_yield == 0)
        )
        finished_df = finished_df[~all_empty_mask].copy()

        finished_df = finished_df[
            finished_df['原料号'].astype(str).str.startswith('3900')
            & finished_df['物料号'].astype(str).str.match(r'^[A-Za-z]')
        ].copy()
        part_order_map = {'胸肉': 0, '腿肉': 1}
        finished_df['_part_order'] = finished_df['部位'].apply(
            lambda x: part_order_map.get(str(x).strip(), 2) if pd.notna(x) else 2
        )
        finished_df = finished_df.sort_values(['_part_order', '物料号', '原料号']).reset_index(drop=True)
        finished_df = finished_df.drop(columns=['_part_order'], errors='ignore')
        ordered = [
            '物料号', '物料描述', '入库数量', '原料号', '原料描述',
            '实际数量', '实际金额', '配方数量',
            '系列', '实际收得率', '配方收得率', '部位', '月份',
        ]
        finished_df = finished_df[[c for c in ordered if c in finished_df.columns]]

    # Keep all materials/rows; do not filter by raw list
    df = df.copy()

    # Build material -> category mapping from raw list (by part)
    mapping = {}
    for mat in df['物料号'].unique():
        part = raw_map.get(mat, '')
        mapping[mat] = _category_from_part(part)

    df = df[df['物料号'].isin(mapping.keys())].copy()

    agg = {}
    for _, row in df.iterrows():
        mat = row['物料号']
        if mat not in agg:
            agg[mat] = {
                '物料号': mat,
                '分类': mapping[mat],
                '物料描述': row['物料描述'],
                '入库数量': 0.0,
                '入库金额': 0.0,
                '调整后实际量': 0.0,
                '调整后实际额': 0.0,
                '碎肉量': 0.0,
                '人工费用实际额': 0.0,
            }
        if row['物料描述']:
            agg[mat]['物料描述'] = row['物料描述']

        raw = row['原料号']
        raw_desc = row['原料描述']

        if raw == '' or pd.isna(raw):
            agg[mat]['入库数量'] += row['入库数量']
            agg[mat]['入库金额'] += row['入库金额']
            continue

        if raw == '人工费用' or ('人工' in raw_desc):
            agg[mat]['人工费用实际额'] += row['实际金额']
            continue

        if row['实际数量'] < 0:
            agg[mat]['碎肉量'] += row['实际数量']
            continue

        agg[mat]['调整后实际量'] += row['实际数量']
        agg[mat]['调整后实际额'] += row['实际金额']

    # Fallback for 入库数量:
    # Primary logic uses main/header rows (原料号为空). Some source files may not carry that row clearly.
    # In that case, align with reference BB2 detail behavior by taking max positive 入库数量 per 物料号.
    try:
        in_qty_fallback = (
            df.assign(_inq=pd.to_numeric(df['入库数量'], errors='coerce').fillna(0.0))
              .groupby('物料号', as_index=False)['_inq']
              .max()
        )
        in_qty_map = dict(zip(in_qty_fallback['物料号'], in_qty_fallback['_inq']))
        for mat in agg:
            if float(agg[mat].get('入库数量', 0.0) or 0.0) == 0.0:
                agg[mat]['入库数量'] = float(in_qty_map.get(mat, 0.0) or 0.0)
    except Exception:
        pass

    # 每个半成品的 BOM（完全对齐参考口径）：
    # 1) BOM描述：等价 XLOOKUP(物料号, XX腿肉/胸肉!S:S, XX腿肉/胸肉!R:R) 的首匹配
    # 2) BOM占比：等价 SUMIFS(XX腿肉/胸肉!U:U, XX腿肉/胸肉!A:A, 物料号)
    #    其中 U 列按模型表定义为“人工费用单耗”
    bom_info = {}
    bom_src = df[df['物料号'].astype(str).str.match(r'^390')].copy()
    if not bom_src.empty:
        for mat, sub in bom_src.groupby('物料号'):
            sub = sub.copy()

            # BOM描述：首个“配方数量>0”的配方文本
            sub_pos = sub[pd.to_numeric(sub['配方数量'], errors='coerce').fillna(0) > 0]
            bom_desc = ''
            if not sub_pos.empty:
                first_desc = sub_pos['原料描述'].dropna().astype(str)
                first_desc = first_desc[first_desc.str.strip() != '']
                if not first_desc.empty:
                    bom_desc = first_desc.iloc[0].strip()

            # BOM占比：SUMIFS(U:U, A:A, 物料号) 等价口径
            # 参考表 U 列 = IF(原料号='人工费用', 入库数量/T, 0)
            # 其中 T 来自 IF(H>0,H,0) 的汇总，即“正向配方数量汇总”
            labor_in = float(agg.get(mat, {}).get('入库数量', 0.0))
            formula_total = (
                pd.to_numeric(sub['配方数量'], errors='coerce')
                .fillna(0)
                .where(pd.to_numeric(sub['配方数量'], errors='coerce').fillna(0) > 0, 0)
                .sum()
            )
            bom_pct = (labor_in / formula_total) if formula_total else 0.0

            bom_info[mat] = {
                'BOM': bom_desc,
                'BOM占比': bom_pct,
                '配方汇总量': float(formula_total),
            }

    # Row order for all line types
    row_order = {
        f'{month_label}实际单价': 0,
        f'{quarter_label}实际单价': 1,
        f'{month_label}规格占比': 2,
        f'{quarter_label}规格占比': 3,
        '差异': 4,
        '对半成品成本的影响-单位成本': 5,
        '对半成品成本的影响-总成本': 6,
    }

    def _to_instock_k_no_dec(in_qty):
        if in_qty in (None, ''):
            return None
        try:
            v = float(in_qty)
        except Exception:
            return None
        if v == 0:
            return None
        # Keep raw precision; display format controls visible decimals.
        return v / 1000.0

    records = []
    for mat, v in agg.items():
        adj_qty = v['调整后实际量']
        adj_amt = v['调整后实际额']
        in_qty = v['入库数量']
        in_qty_k = _to_instock_k_no_dec(in_qty)
        scrap_qty = v['碎肉量']

        auxiliary = adj_qty
        unit = adj_amt / adj_qty if adj_qty != 0 else None
        scrap_ratio = (abs(scrap_qty) / auxiliary) if auxiliary != 0 else None
        util = (in_qty / auxiliary) if auxiliary != 0 else None
        loss = (1 - util - scrap_ratio) if auxiliary != 0 else None
        factor = 0.7 if v['分类'] == '胸肉' else 0.95
        raw_cost = None
        if v['分类'] == '其他':
            raw_cost = unit
        elif unit is not None and util not in (None, 0):
            raw_cost = (unit - (1 - util - loss) * unit * factor) / util
        labor = (v['人工费用实际额'] / in_qty) if in_qty != 0 else None
        total = (raw_cost + labor) if (raw_cost is not None and labor is not None) else None

        records.append({
            '物料号': v['物料号'],
            '分类': v['分类'],
            '物料描述': v['物料描述'],
            '行类型': f'{month_label}实际单价',
            '影响口径': '',
            '行序': row_order[f'{month_label}实际单价'],
            '半成品入库量': in_qty_k,
            'BOM': bom_info.get(mat, {}).get('BOM', ''),
            'BOM占比': bom_info.get(mat, {}).get('BOM占比'),
            '调整后实际量': adj_qty,
            '调整后实际额': adj_amt,
            '碎肉量': scrap_qty,
            '修形前原料综合耗用单价': unit,
            '修形利用率': util,
            '损耗率': loss,
            '半成品原料成本': raw_cost,
            '半成品修形人工成本': labor,
            '半成品总成本': total,
        })

    result = pd.DataFrame(records)
    # Add Q3 actual price / 当月规格占比 / Q3规格占比 rows
    extra_rows = []
    for mat, v in agg.items():
        tsc_sheet = '腿肉TSC' if v['分类'] == '腿肉' else '胸肉TSC'
        q3_ratio = _find_tsc_value(q3_file, tsc_sheet, mat, f'{quarter_label}规格占比')
        nov_ratio = None
        q3_metrics = _find_tsc_metrics(q3_file, tsc_sheet, mat, f'{quarter_label}实际单价')
        in_qty = v.get('入库数量', 0.0)
        in_qty_k = _to_instock_k_no_dec(in_qty)

        for label, val in [
            (f'{month_label}规格占比', nov_ratio),
            (f'{quarter_label}规格占比', q3_ratio),
        ]:
            extra_rows.append({
                '物料号': mat,
                '分类': v['分类'],
                '物料描述': v['物料描述'],
                '行类型': label,
                '影响口径': '',
                '行序': row_order[label],
                '半成品入库量': in_qty_k,
                'BOM': bom_info.get(mat, {}).get('BOM', ''),
                'BOM占比': bom_info.get(mat, {}).get('BOM占比'),
                '调整后实际量': None,
                '调整后实际额': None,
                '碎肉量': None,
                '修形前原料综合耗用单价': val,
                '修形利用率': None,
                '损耗率': None,
                '半成品原料成本': None,
                '半成品修形人工成本': None,
                '半成品总成本': None,
            })

        if q3_metrics:
            extra_rows.append({
                '物料号': mat,
                '分类': v['分类'],
                '物料描述': v['物料描述'],
                '行类型': f'{quarter_label}实际单价',
                '影响口径': '',
                '行序': row_order[f'{quarter_label}实际单价'],
                '半成品入库量': in_qty_k,
                'BOM': bom_info.get(mat, {}).get('BOM', ''),
                'BOM占比': bom_info.get(mat, {}).get('BOM占比'),
                '调整后实际量': None,
                '调整后实际额': None,
                '碎肉量': None,
                '修形前原料综合耗用单价': q3_metrics.get('修形前原料综合耗用单价'),
                '修形利用率': q3_metrics.get('修形利用率'),
                '损耗率': q3_metrics.get('损耗率'),
                '半成品原料成本': q3_metrics.get('半成品原料成本'),
                '半成品修形人工成本': q3_metrics.get('半成品修形人工成本'),
                '半成品总成本': q3_metrics.get('半成品总成本'),
            })

            # 差异 + 对半成品成本的影响（单位成本/总成本）
            in_qty = v.get('入库数量', 0.0)
            in_qty_k = _to_instock_k_no_dec(in_qty)
            adj_qty = v.get('调整后实际量', 0.0)
            adj_amt = v.get('调整后实际额', 0.0)
            scrap_qty = v.get('碎肉量', 0.0)
            month_unit = (adj_amt / adj_qty) if adj_qty else None
            month_util = (in_qty / adj_qty) if adj_qty else None
            month_scrap = (abs(scrap_qty) / adj_qty) if adj_qty else None
            month_loss = (1 - month_util - month_scrap) if adj_qty else None
            factor = 0.7 if v['分类'] == '胸肉' else 0.95
            month_raw_cost = None
            if v['分类'] == '其他':
                month_raw_cost = month_unit
            elif month_unit is not None and month_util not in (None, 0):
                month_raw_cost = (month_unit - (1 - month_util - month_loss) * month_unit * factor) / month_util
            month_labor = (v['人工费用实际额'] / in_qty) if in_qty else None
            month_total = (month_raw_cost + month_labor) if (month_raw_cost is not None and month_labor is not None) else None

            diff_unit = (month_unit - q3_metrics.get('修形前原料综合耗用单价')) if (month_unit is not None and q3_metrics.get('修形前原料综合耗用单价') is not None) else None
            diff_util = (month_util - q3_metrics.get('修形利用率')) if (month_util is not None and q3_metrics.get('修形利用率') is not None) else None
            diff_loss = (month_loss - q3_metrics.get('损耗率')) if (month_loss is not None and q3_metrics.get('损耗率') is not None) else None
            diff_raw_cost = (month_raw_cost - q3_metrics.get('半成品原料成本')) if (month_raw_cost is not None and q3_metrics.get('半成品原料成本') is not None) else None
            diff_labor = (month_labor - q3_metrics.get('半成品修形人工成本')) if (month_labor is not None and q3_metrics.get('半成品修形人工成本') is not None) else None
            diff_total = (month_total - q3_metrics.get('半成品总成本')) if (month_total is not None and q3_metrics.get('半成品总成本') is not None) else None
            diff_ratio = (nov_ratio - q3_ratio) if (nov_ratio is not None and q3_ratio is not None) else None

            def _raw_cost_by(unit_price, util_rate, loss_rate, k):
                if unit_price is None or util_rate in (None, 0) or loss_rate is None:
                    return None
                return (unit_price - (1 - util_rate - loss_rate) * unit_price * k) / util_rate

            extra_rows.append({
                '物料号': mat,
                '分类': v['分类'],
                '物料描述': v['物料描述'],
                '行类型': '差异',
                '影响口径': '',
                '行序': row_order['差异'],
                '半成品入库量': in_qty_k,
                'BOM': bom_info.get(mat, {}).get('BOM', ''),
                'BOM占比': bom_info.get(mat, {}).get('BOM占比'),
                '调整后实际量': None,
                '调整后实际额': None,
                '碎肉量': None,
                # 参考口径：差异首列=当月单价-季度单价
                '修形前原料综合耗用单价': diff_unit,
                '修形利用率': diff_util,
                '损耗率': diff_loss,
                '半成品原料成本': diff_raw_cost,
                '半成品修形人工成本': diff_labor,
                '半成品总成本': diff_total,
            })

            # 参考表“对半成品成本的影响-单位成本”分解口径
            # 1) 单价影响：S(月) - RawCost(Q3单价, 月利用率, 月损耗率)
            unit_impact_price = None
            if diff_unit not in (None, 0):
                rc = _raw_cost_by(
                    q3_metrics.get('修形前原料综合耗用单价'),
                    month_util,
                    month_loss,
                    factor,
                )
                if month_raw_cost is not None and rc is not None:
                    unit_impact_price = month_raw_cost - rc
                else:
                    unit_impact_price = 0

            # 2) 损耗率影响：S(月) - RawCost(月单价, 月利用率, Q3损耗率)
            #    注意：这里严格使用原始实数口径（如 -0.00542053873414），不使用任何显示层四舍五入结果。
            unit_impact_loss = None
            q3_loss_raw = q3_metrics.get('损耗率')
            if month_unit is not None and month_util not in (None, 0) and q3_loss_raw is not None:
                rc = _raw_cost_by(month_unit, month_util, q3_loss_raw, factor)
                if month_raw_cost is not None and rc is not None:
                    unit_impact_loss = month_raw_cost - rc
                else:
                    unit_impact_loss = 0

            # 3) 利用率影响：半成品原料成本差异 - 单价影响 - 损耗率影响
            unit_impact_util = None
            if diff_util not in (None, 0):
                if diff_raw_cost is not None and unit_impact_price is not None and unit_impact_loss is not None:
                    unit_impact_util = diff_raw_cost - unit_impact_price - unit_impact_loss
                else:
                    unit_impact_util = 0

            unit_impact_raw_total = None
            if (
                unit_impact_price is not None
                and unit_impact_util is not None
                and unit_impact_loss is not None
            ):
                unit_impact_raw_total = unit_impact_price + unit_impact_util + unit_impact_loss

            unit_impact_labor = diff_labor
            unit_impact_total = None
            if unit_impact_raw_total is not None and unit_impact_labor is not None:
                unit_impact_total = unit_impact_raw_total + unit_impact_labor

            extra_rows.append({
                '物料号': mat,
                '分类': v['分类'],
                '物料描述': v['物料描述'],
                '行类型': '对半成品成本的影响',
                '影响口径': '单位成本',
                '行序': row_order['对半成品成本的影响-单位成本'],
                '半成品入库量': in_qty_k,
                'BOM': bom_info.get(mat, {}).get('BOM', ''),
                'BOM占比': bom_info.get(mat, {}).get('BOM占比'),
                '调整后实际量': None,
                '调整后实际额': None,
                '碎肉量': None,
                # 参考公式分解
                '修形前原料综合耗用单价': unit_impact_price,
                '修形利用率': unit_impact_util,
                '损耗率': unit_impact_loss,
                '半成品原料成本': unit_impact_raw_total,
                '半成品修形人工成本': unit_impact_labor,
                '半成品总成本': unit_impact_total,
            })
            extra_rows.append({
                '物料号': mat,
                '分类': v['分类'],
                '物料描述': v['物料描述'],
                '行类型': '对半成品成本的影响',
                '影响口径': '总成本',
                '行序': row_order['对半成品成本的影响-总成本'],
                '半成品入库量': in_qty_k,
                'BOM': bom_info.get(mat, {}).get('BOM', ''),
                'BOM占比': bom_info.get(mat, {}).get('BOM占比'),
                '调整后实际量': None,
                '调整后实际额': None,
                '碎肉量': None,
                # 总成本影响口径：单位影响 * 半成品入库量(千kg)
                '修形前原料综合耗用单价': (unit_impact_price * in_qty_k) if (unit_impact_price is not None and in_qty_k is not None) else None,
                '修形利用率': (unit_impact_util * in_qty_k) if (unit_impact_util is not None and in_qty_k is not None) else None,
                '损耗率': (unit_impact_loss * in_qty_k) if (unit_impact_loss is not None and in_qty_k is not None) else None,
                '半成品原料成本': (unit_impact_raw_total * in_qty_k) if (unit_impact_raw_total is not None and in_qty_k is not None) else None,
                '半成品修形人工成本': (unit_impact_labor * in_qty_k) if (unit_impact_labor is not None and in_qty_k is not None) else None,
                '半成品总成本': (unit_impact_total * in_qty_k) if (unit_impact_total is not None and in_qty_k is not None) else None,
            })

    if '行序' not in result.columns:
        result['行序'] = row_order[f'{month_label}实际单价']
    if extra_rows:
        extra_df = pd.DataFrame(extra_rows)
        if not extra_df.dropna(how='all').empty:
            result = pd.concat([result, extra_df], ignore_index=True)
    if '物料号' not in result.columns:
        result['物料号'] = ''
    if '分类' not in result.columns:
        result['分类'] = ''
    if '行序' not in result.columns:
        result['行序'] = 0
    legs = result[result['分类'] == '腿肉'].sort_values(['物料号', '行序'])
    breast = result[result['分类'] == '胸肉'].sort_values(['物料号', '行序'])
    other = result[
        (result['分类'] == '其他') & (result['物料号'].astype(str).str.startswith('390'))
    ].sort_values(['物料号', '行序'])

    # Build raw usage matrix (like TSC F:O)
    raw_cols_legs, comp_col_legs, df_q3_legs, spec_map_legs = _get_tsc_raw_columns(q3_file, '腿肉TSC')
    raw_cols_breast, comp_col_breast, df_q3_breast, spec_map_breast = _get_tsc_raw_columns(q3_file, '胸肉TSC')

    df_calc = df.copy()
    df_calc = df_calc[df_calc['原料号'] != '']
    df_calc['正向数量'] = df_calc['实际数量'].where(df_calc['实际数量'] > 0, 0)
    df_calc['正向金额'] = df_calc['实际金额'].where(df_calc['实际数量'] > 0, 0)

    grp = (
        df_calc.groupby(['物料号', '原料号'], as_index=False)[['正向数量', '正向金额']]
        .sum()
        .rename(columns={'原料号': '原料号_raw'})
    )

    aux_map = {k: v['调整后实际量'] for k, v in agg.items()}
    desc_map = {k: v['物料描述'] for k, v in agg.items()}

    def build_matrix_rows(materials, raw_cols, comp_col, df_q3, category_label):
        rows = []
        base_cols = ['物料号', '物料描述', '行类型', '分类']
        raw_codes = [code for _, code in raw_cols]
        for mat in materials:
            aux = aux_map.get(mat, 0)
            mat_desc = desc_map.get(mat, '')
            mat_rows = grp[grp['物料号'] == mat]
            raw_unit = {}
            raw_ratio = {}
            for _, r in mat_rows.iterrows():
                code = _normalize_mat(r['原料号_raw'])
                qty = r['正向数量']
                amt = r['正向金额']
                raw_unit[code] = (amt / qty) if qty != 0 else None
                raw_ratio[code] = (qty / aux) if aux != 0 else None

            q3_price = _find_tsc_row_values(df_q3, mat, f'{quarter_label}实际单价', raw_cols, comp_col)
            q3_ratio = _find_tsc_row_values(df_q3, mat, f'{quarter_label}规格占比', raw_cols, comp_col)

            for label, source in [
                (f'{month_label}实际单价', raw_unit),
                (f'{quarter_label}实际单价', q3_price),
                (f'{month_label}规格占比', raw_ratio),
                (f'{quarter_label}规格占比', q3_ratio),
            ]:
                row = {'物料号': mat, '物料描述': mat_desc, '行类型': label, '分类': category_label}
                for _, code in raw_cols:
                    row[code] = source.get(code)
                row['综合单价'] = source.get('综合单价')
                rows.append(row)
        if not rows:
            return pd.DataFrame(columns=base_cols + raw_codes + ['综合单价'])
        return pd.DataFrame(rows)

    raw_usage_legs = build_matrix_rows(legs['物料号'].unique(), raw_cols_legs, comp_col_legs, df_q3_legs, '腿肉')
    raw_usage_breast = build_matrix_rows(breast['物料号'].unique(), raw_cols_breast, comp_col_breast, df_q3_breast, '胸肉')

    # Build BB2-style detail sheets
    totals = {}
    for mat, v in agg.items():
        aux = v['调整后实际量']
        unit = v['调整后实际额'] / aux if aux != 0 else 0
        scrap_ratio = abs(v['碎肉量']) / aux if aux != 0 else 0
        totals[mat] = {
            '辅助': aux,
            '总单价': unit,
            '碎肉占比': scrap_ratio,
            '修形利用率': (v['入库数量'] / aux) if aux != 0 else 0,
            '失水率': (1 - (v['入库数量'] / aux) - scrap_ratio) if aux != 0 else 0,
        }

    bb2_rows = []
    for _, row in df.iterrows():
        mat = row['物料号']
        if mat not in totals:
            continue
        total = totals[mat]
        raw = row['原料号']
        raw_desc = row['原料描述']
        is_header = raw == ''
        is_labor = raw == '人工费用' or ('人工' in raw_desc)
        is_scrap = row['实际数量'] < 0

        adj_qty = (
            row['实际数量'] if (not is_header and not is_labor and not is_scrap)
            else (total['辅助'] if is_labor else 0)
        )
        adj_amt = (
            row['实际金额'] if (not is_header and not is_labor and not is_scrap)
            else (agg[mat]['调整后实际额'] if is_labor else 0)
        )
        scrap_qty = row['实际数量'] if is_scrap else (agg[mat]['碎肉量'] if is_labor else 0)

        bb2_rows.append({
            '物料号': mat,
            '分类': mapping.get(mat, '其他'),
            '物料描述(不含琵琶腿/全腿和无抗）': row['物料描述'],
            '入库数量': row['入库数量'] if is_header else (agg[mat]['入库数量'] if is_labor else 0),
            '原料号': raw,
            '原料描述': raw_desc,
            '实际数量': row['实际数量'] if not is_header else 0,
            '实际金额': row['实际金额'] if not is_header else 0,
            '配方数量': row['配方数量'] if not is_header else 0,
            '调整后实际量': adj_qty,
            '辅助': total['辅助'],
            '调整后实际额': adj_amt,
            '碎肉量': scrap_qty,
            '修形前原料占比': (adj_qty / total['辅助']) if total['辅助'] != 0 else 0,
            '修形前原料单价': (adj_amt / adj_qty) if adj_qty != 0 else 0,
            '碎肉占比': (-scrap_qty / total['辅助']) if (is_scrap and total['辅助'] != 0) else (total['碎肉占比'] if is_labor else 0),
            '修形利用率': total['修形利用率'] if is_labor else 0,
            '失水率': total['失水率'] if is_labor else 0,
            '配方': raw_desc if row['配方数量'] > 0 else '',
            '配方物料号': mat if row['配方数量'] > 0 else '',
            '配方数量显示': row['配方数量'] if row['配方数量'] > 0 else 0,
            '配方人工单价': (row['入库数量'] / row['配方数量']) if (raw == '人工费用' and row['配方数量'] > 0) else 0,
            '配方月份': month_code,
        })

    bb2 = pd.DataFrame(bb2_rows)
    if '物料号' not in bb2.columns:
        bb2 = pd.DataFrame(columns=['物料号'])
    bb2 = bb2[bb2['物料号'].astype(str).str.startswith('3900')].copy()
    order_map = {'胸肉': 0, '腿肉': 1, '其他': 2}
    bb2['分类序'] = bb2['分类'].map(order_map).fillna(2)
    bb2 = bb2.sort_values(['分类序', '物料号'])
    bb2 = bb2.drop(columns=['分类序'])
    if '分类' in bb2.columns:
        cols = ['分类'] + [c for c in bb2.columns if c != '分类']
        bb2 = bb2[cols]
    bb2_legs = bb2[bb2['物料号'].isin(legs['物料号'])].copy()
    bb2_breast = bb2[bb2['物料号'].isin(breast['物料号'])].copy()

    labor_records = []
    for mat, v in agg.items():
        mat_norm = _normalize_mat(v.get('物料号', ''))
        if not str(mat_norm).startswith('3900'):
            continue
        amt = v.get('人工费用实际额', 0.0)
        in_qty = v.get('入库数量', 0.0)
        if amt == 0 and in_qty == 0:
            continue
        labor_records.append(
            {
                '分类': v.get('分类', ''),
                '物料号': v.get('物料号', ''),
                '物料描述': v.get('物料描述', ''),
                '入库数量': in_qty,
                '原料号': '人工费用',
                '原料描述': '',
                '实际数量': None,
                '实际金额': amt,
                '人工单耗': (amt / in_qty) if in_qty else None,
            }
        )
    labor_df = pd.DataFrame(labor_records)
    if not labor_df.empty:
        labor_df = labor_df.sort_values(['分类', '物料号']).reset_index(drop=True)
        total_in = labor_df['入库数量'].sum()
        total_amt = labor_df['实际金额'].sum()
        total_row = pd.DataFrame(
            [
                {
                    '分类': '合计',
                    '物料号': '',
                    '物料描述': '',
                    '入库数量': total_in,
                    '原料号': '人工费用',
                    '原料描述': '',
                    '实际数量': None,
                    '实际金额': total_amt,
                    '人工单耗': (total_amt / total_in) if total_in else None,
                }
            ]
        )
        labor_df = pd.concat([total_row, labor_df], ignore_index=True)

    return (
        legs,
        breast,
        other,
        bb2,
        bb2_legs,
        bb2_breast,
        raw_usage_legs,
        raw_usage_breast,
        spec_map_legs,
        spec_map_breast,
        finished_df,
        labor_df,
        semi_category,
    )


def to_excel_bytes(
    legs,
    breast,
    bb2_all,
    bb2_legs,
    bb2_breast,
    raw_usage_legs,
    raw_usage_breast,
    spec_map_legs,
    spec_map_breast,
    prefix,
    finished_df,
    labor_df,
    month_label,
    semi_category,
):
    output = io.BytesIO()
    fmt_money = '{:,.2f}'
    fmt_pct = '{:.0%}'
    fmt_int = '{:,.0f}'

    def _is_num(x):
        return isinstance(x, (int, float)) and not isinstance(x, bool)

    def _fmt_or_dash(x, fmt):
        if x is None or x == '':
            return ''
        if _is_num(x):
            if pd.isna(x):
                return '-'
            return '-' if x == 0 else fmt.format(x)
        return '-' if str(x).lower() == 'nan%' else x

    def _fmt_keep_zero(x, fmt):
        if x is None or x == '':
            return ''
        if _is_num(x):
            if pd.isna(x):
                return '-'
            return fmt.format(x)
        return '-' if str(x).lower() == 'nan%' else x

    def _fmt_pct_or_dash(x):
        if x is None or x == '':
            return ''
        if _is_num(x):
            if pd.isna(x) or x == 0:
                return '-'
            pct = f'{abs(x):.0%}'
            return f'({pct})' if x < 0 else pct
        sx = str(x).strip()
        if sx.lower() == 'nan%' or sx == '0%':
            return '-'
        return sx

    def _fmt_pct1_or_dash(x):
        if x is None or x == '':
            return ''
        if _is_num(x):
            if pd.isna(x):
                return '-'
            return f'{x:.1%}'
        sx = str(x).strip()
        if sx.lower() == 'nan%' or sx == '0%':
            return '-'
        return sx

    def _numeric_text_to_number(x):
        # Keep non-numeric ids (e.g. CF..., 人工费用) unchanged; only convert pure digit text.
        if x is None:
            return x
        s = str(x).strip()
        if s in ('', '-'):
            return x
        if re.fullmatch(r'\d+', s):
            try:
                return int(s)
            except Exception:
                return x
        return x

    def apply_formats(df):
        df = df.copy()
        df = df.drop(columns=[c for c in HIDE_COLS if c in df.columns])
        df = df.drop(columns=['行序'], errors='ignore')
        impact_mask = pd.Series(False, index=df.index)
        total_impact_mask = pd.Series(False, index=df.index)
        diff_mask = pd.Series(False, index=df.index)
        if '行类型' in df.columns:
            impact_mask = df['行类型'].astype(str).eq('对半成品成本的影响')
            diff_mask = df['行类型'].astype(str).eq('差异')
        if impact_mask.any() and '影响口径' in df.columns:
            total_impact_mask = impact_mask & df['影响口径'].astype(str).eq('总成本')
        show_actual_mask = impact_mask | diff_mask

        for col in ['修形前原料综合耗用单价', '半成品原料成本', '半成品修形人工成本', '半成品总成本']:
            if col in df.columns:
                # 常规行与单位成本影响：两位小数；总成本影响：整数
                df.loc[~show_actual_mask & ~total_impact_mask, col] = df.loc[~show_actual_mask & ~total_impact_mask, col].apply(lambda x: _fmt_or_dash(x, fmt_money))
                df.loc[(show_actual_mask | total_impact_mask) & ~total_impact_mask, col] = df.loc[(show_actual_mask | total_impact_mask) & ~total_impact_mask, col].apply(lambda x: _fmt_keep_zero(x, fmt_money))
                df.loc[total_impact_mask, col] = df.loc[total_impact_mask, col].apply(lambda x: _fmt_keep_zero(x, fmt_int))
        for col in ['修形利用率', '损耗率']:
            if col in df.columns:
                # 常规行：百分比；影响行改为数值。
                # 注意：损耗率-差异必须直接基于原始实数做 1 位小数百分比，不能先被 0 位百分比覆盖
                if col == '损耗率':
                    df.loc[~impact_mask & ~diff_mask, col] = df.loc[~impact_mask & ~diff_mask, col].apply(_fmt_pct_or_dash)
                    df.loc[diff_mask, col] = df.loc[diff_mask, col].apply(_fmt_pct1_or_dash)
                else:
                    df.loc[~impact_mask, col] = df.loc[~impact_mask, col].apply(_fmt_pct_or_dash)
                df.loc[impact_mask & ~total_impact_mask, col] = df.loc[impact_mask & ~total_impact_mask, col].apply(lambda x: _fmt_keep_zero(x, fmt_money))
                if col == '损耗率':
                    # 总成本-损耗率按两位小数展示（如 50.15）
                    df.loc[total_impact_mask, col] = df.loc[total_impact_mask, col].apply(lambda x: _fmt_keep_zero(x, fmt_money))
                else:
                    df.loc[total_impact_mask, col] = df.loc[total_impact_mask, col].apply(lambda x: _fmt_keep_zero(x, fmt_int))
        return df

    def apply_bb2_formats(df):
        df = df.copy()
        money_cols = ['入库金额', '配方人工单价']
        money_int_cols = ['实际金额', '调整后实际额']
        qty_cols = ['入库数量', '实际数量', '配方数量', '调整后实际量', '辅助', '碎肉量']
        qty_cols += ['配方数量显示']
        unit_cols = ['修形前原料单价', '总单价', '人工单耗']
        ratio_cols = ['修形前原料占比', '碎肉占比', '修形利用率', '失水率']
        decimal_ratio_cols = ['实际收得率', '配方收得率']

        # Keep numeric columns as true numbers (avoid Excel "number stored as text" warnings)
        for col in money_cols + money_int_cols + qty_cols + unit_cols + ratio_cols + decimal_ratio_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        # Avoid "number stored as text" warnings in Excel for id columns with pure digits
        for col in ['物料号', '原料号', '配方物料号']:
            if col in df.columns:
                df[col] = df[col].apply(_numeric_text_to_number)
        return df

    def _build_semi_summary(df):
        if df.empty or '物料号' not in df.columns:
            return pd.DataFrame(columns=['物料号', '物料描述', '入库数量', '系列'])
        desc_col = '物料描述(不含琵琶腿/全腿和无抗）'
        if desc_col not in df.columns:
            desc_col = '物料描述'
        grp = df.groupby('物料号', as_index=False).agg(
            物料描述=(desc_col, lambda x: next((v for v in x if str(v).strip() != ''), '')),
            入库数量=('入库数量', 'max'),
        )
        grp['系列'] = grp['物料号'].apply(
            lambda x: semi_category.get(_normalize_mat(x), '')
            if isinstance(semi_category, dict)
            else ''
        )
        return grp[['物料号', '物料描述', '入库数量', '系列']]

    def _fmt_int_or_dash(x):
        if x is None or x == '':
            return ''
        if isinstance(x, (int, float)) and not isinstance(x, bool):
            if pd.isna(x):
                return '-'
            return '-' if x == 0 else fmt_int.format(x)
        return x

    # Keep raw numeric values for TSC export (no pre-rounding/string formatting).
    legs_raw_for_tsc = legs.copy()
    breast_raw_for_tsc = breast.copy()
    # Display/export of non-TSC sections can keep formatted style.
    legs = apply_formats(legs)
    breast = apply_formats(breast)

    def build_tsc_sheet(tsc_df, raw_usage_df, spec_map):
        base_cols = ['物料号', '物料描述', '行类型', '影响口径', '分类']
        raw_cols = [c for c in raw_usage_df.columns if c not in base_cols + ['综合单价']]
        raw_cols = [c for c in raw_cols if c != '综合单价']

        merged = tsc_df.merge(
            raw_usage_df,
            on=['物料号', '物料描述', '行类型'],
            how='left',
            suffixes=('', '_raw'),
        )
        # Ensure 半成品入库量 always exists in TSC and can be positioned before BOM.
        if '半成品入库量' not in merged.columns:
            in_qty_map = (
                tsc_df[['物料号', '半成品入库量']]
                .dropna(subset=['物料号'])
                .drop_duplicates(subset=['物料号'])
            ) if '半成品入库量' in tsc_df.columns else pd.DataFrame(columns=['物料号', '半成品入库量'])
            if not in_qty_map.empty:
                merged = merged.merge(in_qty_map, on='物料号', how='left')
            else:
                merged['半成品入库量'] = None
        merged['修行后原料'] = merged['物料号']
        merged['使用半成品规格'] = merged['物料描述']
        if '修行后原料' in merged.columns:
            merged['产品族'] = merged['修行后原料'].apply(_normalize_mat).map(semi_category).fillna('')
        else:
            merged['产品族'] = ''

        ordered_cols = (
            ['产品族', '修行后原料', '使用半成品规格', '行类型', '影响口径']
            + raw_cols
            + ['综合单价']
            + ['修形前原料综合耗用单价', '修形利用率', '损耗率', '半成品原料成本', '半成品修形人工成本', '半成品总成本']
            + ['半成品入库量', 'BOM', 'BOM占比']
        )
        # Keep only columns that exist
        ordered_cols = [c for c in ordered_cols if c in merged.columns]
        header1 = {c: c for c in ordered_cols}
        header2 = {c: '' for c in ordered_cols}
        for code in raw_cols:
            header2[code] = spec_map.get(code, '')
        if '综合单价' in ordered_cols:
            header2['综合单价'] = '综合单价'
        blank = {c: '' for c in ordered_cols}
        header0 = {c: '' for c in ordered_cols}
        cost_cols = [
            '修形前原料综合耗用单价',
            '修形利用率',
            '损耗率',
            '半成品原料成本',
            '半成品修形人工成本',
            '半成品总成本',
        ]
        for c in cost_cols:
            if c in header0:
                header0[c] = '修形后成本'
        # Keep numeric values as-is; formatting should be applied at worksheet layer.
        merged_vals = merged[ordered_cols].copy()

        # x月实际单价行：综合单价对齐为“修形前原料综合耗用单价”
        if '行类型' in merged_vals.columns and '综合单价' in merged_vals.columns and '修形前原料综合耗用单价' in merged_vals.columns:
            month_price_mask = merged_vals['行类型'].astype(str).eq(f'{month_label}实际单价')
            merged_vals.loc[month_price_mask, '综合单价'] = merged_vals.loc[month_price_mask, '修形前原料综合耗用单价']

        return pd.concat(
            [pd.DataFrame([header0, header1, header2, blank]), merged_vals],
            ignore_index=True,
        )

    tsc_legs = build_tsc_sheet(legs_raw_for_tsc, raw_usage_legs, spec_map_legs)
    tsc_breast = build_tsc_sheet(breast_raw_for_tsc, raw_usage_breast, spec_map_breast)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        apply_bb2_formats(finished_df).to_excel(writer, index=False, sheet_name=f'成品-{month_label}')
        apply_bb2_formats(labor_df).to_excel(writer, index=False, sheet_name='人工')
        apply_bb2_formats(bb2_all).to_excel(writer, index=False, sheet_name='半成品')
        apply_bb2_formats(bb2_legs).to_excel(writer, index=False, sheet_name=f'{prefix}腿肉')
        tsc_legs.to_excel(writer, index=False, sheet_name='腿肉TSC', header=False)
        apply_bb2_formats(bb2_breast).to_excel(writer, index=False, sheet_name=f'{prefix}胸肉')
        tsc_breast.to_excel(writer, index=False, sheet_name='胸肉TSC', header=False)

        header_fill = PatternFill(fill_type='solid', fgColor='FFFF00')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_font = Font(bold=True)

        def _style_tsc_headers(ws, tsc_df):
            # header rows are 1..3 (header0, header1, header2)
            for row in range(1, 4):
                for cell in ws[row]:
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.font = header_font

            cost_cols = [
                '修形前原料综合耗用单价',
                '修形利用率',
                '损耗率',
                '半成品原料成本',
                '半成品修形人工成本',
                '半成品总成本',
            ]
            cols = list(tsc_df.columns)
            cost_idxs = [cols.index(c) + 1 for c in cost_cols if c in cols]
            if cost_idxs:
                ws.merge_cells(
                    start_row=1,
                    start_column=min(cost_idxs),
                    end_row=1,
                    end_column=max(cost_idxs),
                )
            # set specific column widths for readability
            col_widths = {
                '使用半成品规格': 28,
                '行类型': 13,
            }
            for name, width in col_widths.items():
                if name in cols:
                    col_idx = cols.index(name) + 1
                    ws.column_dimensions[get_column_letter(col_idx)].width = width

        def _display_width(v):
            s = str(v)
            w = 0
            for ch in s:
                w += 2 if unicodedata.east_asian_width(ch) in ('F', 'W') else 1
            return w

        def _autofit(ws, min_width=8, max_width=42, padding=2):
            for col_cells in ws.columns:
                max_len = 0
                col_letter = None
                for cell in col_cells:
                    if cell is not None and hasattr(cell, 'column'):
                        col_letter = getattr(cell, 'column_letter', None) or get_column_letter(cell.column)
                        break
                if not col_letter:
                    continue
                for cell in col_cells:
                    val = cell.value
                    if val is None:
                        continue
                    val_len = _display_width(val)
                    if val_len > max_len:
                        max_len = val_len
                width = min(max_len + padding, max_width)
                width = max(width, min_width)
                ws.column_dimensions[col_letter].width = width

        def _style_bb2_headers(ws):
            # color groups like the reference image
            green_fill = PatternFill(fill_type='solid', fgColor='92D050')
            red_fill = PatternFill(fill_type='solid', fgColor='FFC7CE')
            red_font = Font(color='9C0006', bold=True)
            header_font = Font(bold=True)
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            yellow_cols = ['实际数量', '实际金额', '配方数量']
            green_cols = ['调整后实际量', '辅助', '调整后实际额', '碎肉量']
            red_cols = ['修形前原料占比', '修形前原料单价', '碎肉占比', '修形利用率', '失水率']

            headers = [cell.value for cell in ws[1]]
            for idx, name in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=idx)
                cell.alignment = header_align
                if name in yellow_cols:
                    cell.fill = PatternFill(fill_type='solid', fgColor='FFFF00')
                    cell.font = header_font
                elif name in green_cols:
                    cell.fill = green_fill
                    cell.font = header_font
                elif name in red_cols:
                    cell.fill = PatternFill(fill_type=None)
                    cell.font = red_font
                else:
                    cell.font = header_font

        def _align_bb2_sheet(ws):
            headers = [cell.value for cell in ws[1]]
            left_cols = {'物料号', '分类', '物料描述(不含琵琶腿/全腿和无抗）', '原料号', '原料描述', '配方'}
            right_cols = {'入库数量', '入库金额', '实际数量', '实际金额', '配方数量', '调整后实际量', '辅助', '调整后实际额', '碎肉量', '配方数量显示'}
            center_cols = {'修形前原料占比', '修形前原料单价', '碎肉占比', '修形利用率', '失水率', '配方人工单价', '配方月份'}
            for idx, name in enumerate(headers, start=1):
                if name in left_cols:
                    align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                elif name in right_cols:
                    align = Alignment(horizontal='right', vertical='center')
                elif name in center_cols:
                    align = Alignment(horizontal='center', vertical='center')
                else:
                    align = Alignment(horizontal='center', vertical='center')
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=idx).alignment = align
                    if ws.cell(row=r, column=idx).value == '-':
                        ws.cell(row=r, column=idx).alignment = Alignment(horizontal='center', vertical='center')

        def _style_finished_header(ws):
            header_font = Font(bold=True)
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_align

        def _style_header_row(ws, header_row=1):
            header_font = Font(bold=True)
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for cell in ws[header_row]:
                cell.font = header_font
                cell.alignment = header_align

        def _align_finished_sheet(ws, header_row=1, start_data_row=2):
            headers = [cell.value for cell in ws[header_row]]
            left_cols = {'物料号', '物料描述', '原料号', '原料描述'}
            right_cols = {'入库数量', '实际数量', '实际金额', '配方数量', '实际收得率', '配方收得率'}
            center_cols = {'系列', '部位', '月份'}
            for idx, name in enumerate(headers, start=1):
                if name in left_cols:
                    align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                elif name in right_cols:
                    align = Alignment(horizontal='right', vertical='center')
                elif name in center_cols:
                    align = Alignment(horizontal='center', vertical='center')
                else:
                    align = Alignment(horizontal='center', vertical='center')
                for r in range(start_data_row, ws.max_row + 1):
                    ws.cell(row=r, column=idx).alignment = align
                    if ws.cell(row=r, column=idx).value == '-':
                        ws.cell(row=r, column=idx).alignment = Alignment(horizontal='center', vertical='center')

        def _style_finished_empty_part_raw(ws):
            headers = [cell.value for cell in ws[1]]
            if not headers:
                return
            col_idx = {name: i + 1 for i, name in enumerate(headers)}
            part_col = col_idx.get('部位')
            raw_no_col = col_idx.get('原料号')
            raw_desc_col = col_idx.get('原料描述')
            if not part_col or not raw_no_col or not raw_desc_col:
                return
            blue_fill = PatternFill(fill_type='solid', fgColor='00B0F0')
            for r in range(2, ws.max_row + 1):
                part_val = ws.cell(row=r, column=part_col).value
                if part_val is None or str(part_val).strip() in {'', '-'}:
                    ws.cell(row=r, column=raw_no_col).fill = blue_fill
                    ws.cell(row=r, column=raw_desc_col).fill = blue_fill

        def _align_tsc_sheet(ws, tsc_df):
            cols = list(tsc_df.columns)
            left_cols = {'产品族', '修行后原料', '使用半成品规格', '物料号', '物料描述', '影响口径', 'BOM'}
            center_cols = {
                '行类型', '综合单价',
                '修形前原料综合耗用单价', '修形利用率', '损耗率',
                '半成品原料成本', '半成品修形人工成本', '半成品总成本',
                '半成品入库量', 'BOM占比',
            }
            for idx, name in enumerate(cols, start=1):
                if name in left_cols:
                    align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                elif name in center_cols:
                    align = Alignment(horizontal='center', vertical='center')
                else:
                    align = Alignment(horizontal='right', vertical='center')
                for r in range(4, ws.max_row + 1):
                    ws.cell(row=r, column=idx).alignment = align
                    if ws.cell(row=r, column=idx).value == '-':
                        ws.cell(row=r, column=idx).alignment = Alignment(horizontal='center', vertical='center')

        def _apply_tsc_number_formats(ws, tsc_df):
            cols = list(tsc_df.columns)
            if not cols:
                return
            col_idx = {name: i + 1 for i, name in enumerate(cols)}

            fixed_cols = {
                '产品族', '修行后原料', '使用半成品规格', '行类型', '影响口径', '综合单价',
                '修形前原料综合耗用单价', '修形利用率', '损耗率',
                '半成品原料成本', '半成品修形人工成本', '半成品总成本',
                '半成品入库量', 'BOM', 'BOM占比',
            }
            raw_cols = [c for c in cols if c not in fixed_cols]

            fmt_unit = '0.00;[Red](0.00);-'
            fmt_qty = '#,##0;[Red](#,##0);-'
            fmt_pct0 = '0%;[Red](0%);-'
            fmt_pct1 = '0.0%;[Red](0.0%);-'

            row_type_col = col_idx.get('行类型')
            impact_col = col_idx.get('影响口径')
            start_row = 5  # 1~4 are header/blank rows

            for r in range(start_row, ws.max_row + 1):
                row_type = str(ws.cell(r, row_type_col).value).strip() if row_type_col else ''
                impact_type = str(ws.cell(r, impact_col).value).strip() if impact_col else ''

                is_price_row = ('实际单价' in row_type)
                is_ratio_row = ('规格占比' in row_type)
                is_diff_row = (row_type == '差异')
                is_impact_row = (row_type == '对半成品成本的影响')

                # Raw matrix + 综合单价 follows row type.
                target_cols = raw_cols + (['综合单价'] if '综合单价' in col_idx else [])
                for name in target_cols:
                    c = col_idx.get(name)
                    if not c:
                        continue
                    cell = ws.cell(r, c)
                    if not isinstance(cell.value, (int, float)) or isinstance(cell.value, bool):
                        continue
                    if is_price_row:
                        cell.number_format = fmt_unit
                    elif is_ratio_row:
                        cell.number_format = fmt_pct0

                # Cost block
                for name in ['修形前原料综合耗用单价', '半成品原料成本', '半成品修形人工成本', '半成品总成本']:
                    c = col_idx.get(name)
                    if not c:
                        continue
                    cell = ws.cell(r, c)
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.number_format = fmt_unit

                # Utilization / loss keep raw calc value and show as percent.
                c_util = col_idx.get('修形利用率')
                if c_util:
                    cell = ws.cell(r, c_util)
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        if is_impact_row:
                            # 影响行改为数值展示；总成本口径不保留小数。
                            cell.number_format = fmt_qty if impact_type == '总成本' else fmt_unit
                        else:
                            cell.number_format = fmt_pct0

                c_loss = col_idx.get('损耗率')
                if c_loss:
                    cell = ws.cell(r, c_loss)
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        # 影响行（单位/总成本）按两位小数数值显示；其它保持百分比。
                        if is_impact_row:
                            cell.number_format = fmt_unit
                        else:
                            cell.number_format = fmt_pct1 if is_diff_row else fmt_pct0

                c_in = col_idx.get('半成品入库量')
                if c_in:
                    cell = ws.cell(r, c_in)
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.number_format = fmt_qty

                c_bomr = col_idx.get('BOM占比')
                if c_bomr:
                    cell = ws.cell(r, c_bomr)
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.number_format = fmt_pct1

        def _style_other_rows(ws):
            # apply blue fill for rows where 分类 == '其他' (only selected columns)
            blue_fill = PatternFill(fill_type='solid', fgColor='00B0F0')
            headers = [cell.value for cell in ws[1]]
            if '分类' not in headers:
                return
            cat_col = headers.index('分类') + 1
            target_cols = []
            for name in ['物料号', '分类', '物料描述(不含琵琶腿/全腿和无抗）']:
                if name in headers:
                    target_cols.append(headers.index(name) + 1)
            if not target_cols:
                return
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=cat_col).value
                if str(val).strip() == '其他':
                    for c in target_cols:
                        ws.cell(row=r, column=c).fill = blue_fill

        def _style_labor_rows(ws):
            # highlight rows where 原料号 == 人工费用
            yellow_fill = PatternFill(fill_type='solid', fgColor='FFFF00')
            headers = [cell.value for cell in ws[1]]
            if '原料号' not in headers:
                return
            raw_col = headers.index('原料号') + 1
            max_col = ws.max_column
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=raw_col).value
                if str(val).strip() == '人工费用':
                    for c in range(1, max_col + 1):
                        ws.cell(row=r, column=c).fill = yellow_fill

        def _style_labor_other_rows(ws):
            # apply blue fill for rows where 分类 == '其他' in 人工 sheet
            blue_fill = PatternFill(fill_type='solid', fgColor='00B0F0')
            header_row = 1
            headers = [cell.value for cell in ws[header_row]]
            if '分类' not in headers and ws.max_row >= 2:
                header_row = 2
                headers = [cell.value for cell in ws[header_row]]
            if '分类' not in headers:
                return
            cat_col = headers.index('分类') + 1
            target_cols = []
            for name in ['分类', '物料号', '物料描述']:
                if name in headers:
                    target_cols.append(headers.index(name) + 1)
            if not target_cols:
                return
            for r in range(header_row + 1, ws.max_row + 1):
                val = ws.cell(row=r, column=cat_col).value
                if str(val).strip() == '其他':
                    for c in target_cols:
                        ws.cell(row=r, column=c).fill = blue_fill

        def _apply_number_formats(ws, header_row=1, start_data_row=2):
            headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
            qty_cols = {'入库数量', '实际数量', '配方数量', '调整后实际量', '辅助', '碎肉量', '配方数量显示', '调整后耗用量', '耗用量', '半成品入库量'}
            amt_cols = {'入库金额', '实际金额', '调整后实际额', '调整后金额', '耗用金额'}
            unit_cols = {'单价', '修形前原料单价', '修形前原料综合耗用单价', '半成品原料成本', '半成品修形人工成本', '半成品总成本', '综合单价', '配方人工单价', '人工单耗'}
            pct_cols = {'修形前原料占比', '碎肉占比', '修形利用率', '失水率', '损耗率', 'BOM占比'}
            dec_cols = {'实际收得率', '配方收得率'}
            month_cols = {'配方月份', '月份'}

            fmt_num = '#,##0.##;[Red](#,##0.##);-'
            fmt_unit = '0.00;[Red](0.00);-'
            fmt_pct = '0%;[Red]-0%;-'
            fmt_dec = '0.00;[Red](0.00);-'

            for idx, name in enumerate(headers, start=1):
                if name in qty_cols or name in amt_cols:
                    fmt = fmt_num
                elif name in unit_cols:
                    fmt = fmt_unit
                elif name in pct_cols:
                    fmt = fmt_pct
                elif name in dec_cols:
                    fmt = fmt_dec
                elif name in month_cols:
                    fmt = '0'
                else:
                    continue
                for r in range(start_data_row, ws.max_row + 1):
                    cell = ws.cell(row=r, column=idx)
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.number_format = fmt

        def _append_labor_summary(ws, labor_src_df):
            if labor_src_df is None or labor_src_df.empty:
                return
            if '分类' not in labor_src_df.columns:
                return

            src = labor_src_df[labor_src_df['分类'].isin(['胸肉', '腿肉'])].copy()
            if src.empty:
                return

            for c in ['入库数量', '实际金额']:
                if c in src.columns:
                    src[c] = pd.to_numeric(src[c], errors='coerce').fillna(0)
                else:
                    src[c] = 0

            chest_qty = round(float(src.loc[src['分类'] == '胸肉', '入库数量'].sum()), 2)
            leg_qty = round(float(src.loc[src['分类'] == '腿肉', '入库数量'].sum()), 2)
            chest_amt = round(float(src.loc[src['分类'] == '胸肉', '实际金额'].sum()), 2)
            leg_amt = round(float(src.loc[src['分类'] == '腿肉', '实际金额'].sum()), 2)

            chest_ratio = (chest_amt / chest_qty) if chest_qty else None
            leg_ratio = (leg_amt / leg_qty) if leg_qty else None

            start_col = ws.max_column + 2
            start_row = 1
            headers = ['胸肉', '腿肉']
            data_rows = [
                [chest_qty, leg_qty],
                [chest_amt, leg_amt],
                [chest_ratio, leg_ratio],
            ]

            thin = Side(border_style='thin', color='000000')
            cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for i, h in enumerate(headers):
                cell = ws.cell(row=start_row, column=start_col + i, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = cell_border

            for r_idx, row_vals in enumerate(data_rows, start=1):
                for c_idx, v in enumerate(row_vals):
                    cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=v)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = cell_border
                    if r_idx in (1, 2):
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '0.00'

            ws.column_dimensions[get_column_letter(start_col)].width = 12
            ws.column_dimensions[get_column_letter(start_col + 1)].width = 12

        def _fill_desc_blanks(ws):
            # fill blank descriptions by carrying forward within same material
            headers = [cell.value for cell in ws[1]]
            if '物料号' not in headers or '物料描述(不含琵琶腿/全腿和无抗）' not in headers:
                return
            mat_col = headers.index('物料号') + 1
            desc_col = headers.index('物料描述(不含琵琶腿/全腿和无抗）') + 1
            last_desc = {}
            for r in range(2, ws.max_row + 1):
                mat = ws.cell(row=r, column=mat_col).value
                desc_cell = ws.cell(row=r, column=desc_col)
                if mat is None or str(mat).strip() == '':
                    continue
                key = str(mat).strip()
                if desc_cell.value is None or str(desc_cell.value).strip() == '':
                    if key in last_desc:
                        desc_cell.value = last_desc[key]
                else:
                    last_desc[key] = desc_cell.value

        for sheet_name in ['半成品', f'{prefix}腿肉', f'{prefix}胸肉']:
            if sheet_name in writer.book.sheetnames:
                _style_bb2_headers(writer.book[sheet_name])
                _align_bb2_sheet(writer.book[sheet_name])
                _style_labor_rows(writer.book[sheet_name])
                _fill_desc_blanks(writer.book[sheet_name])
                _apply_number_formats(writer.book[sheet_name], header_row=1, start_data_row=2)
        if '半成品' in writer.book.sheetnames:
            _style_other_rows(writer.book['半成品'])

        finished_sheet = f'成品-{month_label}'
        if finished_sheet in writer.book.sheetnames:
            _style_finished_header(writer.book[finished_sheet])
            _align_finished_sheet(writer.book[finished_sheet])
            _style_finished_empty_part_raw(writer.book[finished_sheet])
            _apply_number_formats(writer.book[finished_sheet], header_row=1, start_data_row=2)
            # 成品表：实际收得率/配方收得率表头留空
            ws_fin = writer.book[finished_sheet]
            for c in range(1, ws_fin.max_column + 1):
                v = ws_fin.cell(row=1, column=c).value
                if str(v).strip() in {'实际收得率', '配方收得率'}:
                    ws_fin.cell(row=1, column=c, value='')
        if '人工' in writer.book.sheetnames:
            ws_labor = writer.book['人工']
            # 调整为：第1行合计，第2行表头，第3行起明细
            ws_labor.insert_rows(1)
            for c in range(1, ws_labor.max_column + 1):
                ws_labor.cell(row=1, column=c, value=ws_labor.cell(row=3, column=c).value)
            for c in range(1, ws_labor.max_column + 1):
                ws_labor.cell(row=3, column=c, value=None)
            # 去掉重复“合计”行（保留第1行合计）
            for r in range(3, ws_labor.max_row + 1):
                if str(ws_labor.cell(row=r, column=1).value).strip() == '合计':
                    ws_labor.delete_rows(r, 1)
                    break
            _style_header_row(ws_labor, header_row=2)
            _align_finished_sheet(ws_labor, header_row=2, start_data_row=1)
            _style_labor_other_rows(ws_labor)
            _append_labor_summary(ws_labor, labor_df)
            _apply_number_formats(ws_labor, header_row=2, start_data_row=1)

        _style_tsc_headers(writer.book['腿肉TSC'], tsc_legs)
        _style_tsc_headers(writer.book['胸肉TSC'], tsc_breast)
        _align_tsc_sheet(writer.book['腿肉TSC'], tsc_legs)
        _align_tsc_sheet(writer.book['胸肉TSC'], tsc_breast)
        _apply_tsc_number_formats(writer.book['腿肉TSC'], tsc_legs)
        _apply_tsc_number_formats(writer.book['胸肉TSC'], tsc_breast)

        def _fill_tsc_instock_qty(ws, tsc_df, bb2_df):
            cols = list(tsc_df.columns)
            if '半成品入库量' not in cols or '修行后原料' not in cols:
                return
            qty_col = cols.index('半成品入库量') + 1
            mat_col = cols.index('修行后原料') + 1

            # 1) from tsc_df itself
            qty_map = {}
            try:
                t = tsc_df[['修行后原料', '半成品入库量']].copy()
                t['修行后原料'] = t['修行后原料'].apply(_normalize_mat)
                t['半成品入库量'] = pd.to_numeric(t['半成品入库量'], errors='coerce')
                t = t.dropna(subset=['修行后原料', '半成品入库量'])
                for _, r in t.iterrows():
                    if r['修行后原料'] and r['修行后原料'] not in qty_map:
                        qty_map[r['修行后原料']] = float(r['半成品入库量'])
            except Exception:
                pass

            # 2) fallback from BB2 sheet: 入库数量 / 1000
            if (not qty_map) and bb2_df is not None and not bb2_df.empty:
                try:
                    b = bb2_df[['物料号', '入库数量']].copy()
                    b['物料号'] = b['物料号'].apply(_normalize_mat)
                    b['入库数量'] = pd.to_numeric(b['入库数量'], errors='coerce').fillna(0.0)
                    b = b.groupby('物料号', as_index=False)['入库数量'].max()
                    for _, r in b.iterrows():
                        if r['物料号']:
                            qty_map[r['物料号']] = float(r['入库数量']) / 1000.0
                except Exception:
                    pass

            if not qty_map:
                return

            qty_fmt = '#,##0;[Red](#,##0);-'
            for r in range(5, ws.max_row + 1):
                mat = _normalize_mat(ws.cell(r, mat_col).value)
                if not mat:
                    continue
                cell = ws.cell(r, qty_col)
                if cell.value in (None, '', '-'):
                    v = qty_map.get(mat)
                    if v is not None:
                        cell.value = v
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = qty_fmt
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        _fill_tsc_instock_qty(writer.book['腿肉TSC'], tsc_legs, bb2_legs)
        _fill_tsc_instock_qty(writer.book['胸肉TSC'], tsc_breast, bb2_breast)

        def _merge_same_spec(ws, tsc_df):
            # merge consecutive rows with same 使用半成品规格 (within same 修行后原料)
            cols = list(tsc_df.columns)
            if '使用半成品规格' not in cols or '修行后原料' not in cols:
                return
            spec_col = cols.index('使用半成品规格') + 1
            mat_col = cols.index('修行后原料') + 1
            merge_cols = [spec_col]
            if '半成品入库量' in cols:
                merge_cols.append(cols.index('半成品入库量') + 1)
            if 'BOM' in cols:
                merge_cols.append(cols.index('BOM') + 1)
            if 'BOM占比' in cols:
                merge_cols.append(cols.index('BOM占比') + 1)
            start_row = 5  # data starts after 4 header rows
            current_val = None
            current_mat = None
            merge_start = None

            def _merge_group(r1, r2):
                if r2 - r1 <= 0:
                    return
                for c in merge_cols:
                    # keep first non-empty value at top cell before merge
                    top_val = ws.cell(row=r1, column=c).value
                    if top_val in (None, ''):
                        for rr in range(r1, r2 + 1):
                            vv = ws.cell(row=rr, column=c).value
                            if vv not in (None, ''):
                                top_val = vv
                                break
                    ws.merge_cells(start_row=r1, start_column=c, end_row=r2, end_column=c)
                    ws.cell(row=r1, column=c, value=top_val)

            for r in range(start_row, ws.max_row + 2):
                val = ws.cell(row=r, column=spec_col).value if r <= ws.max_row else None
                mat = ws.cell(row=r, column=mat_col).value if r <= ws.max_row else None
                key_val = (str(val).strip() if val is not None else '')
                key_mat = (str(mat).strip() if mat is not None else '')
                if merge_start is None:
                    current_val, current_mat = key_val, key_mat
                    merge_start = r
                    continue
                if key_val == current_val and key_mat == current_mat and key_val != '':
                    continue
                # close previous group
                if merge_start is not None and r - merge_start > 1 and current_val != '':
                    _merge_group(merge_start, r - 1)
                current_val, current_mat = key_val, key_mat
                merge_start = r
        def _merge_product_group(ws, tsc_df):
            cols = list(tsc_df.columns)
            if '产品族' not in cols or '修行后原料' not in cols:
                return
            prod_col = cols.index('产品族') + 1
            mat_col = cols.index('修行后原料') + 1
            start_row = 5
            current_mat = None
            merge_start = None
            for r in range(start_row, ws.max_row + 2):
                mat = ws.cell(row=r, column=mat_col).value if r <= ws.max_row else None
                key_mat = (str(mat).strip() if mat is not None else '')
                if merge_start is None:
                    current_mat = key_mat
                    merge_start = r
                    continue
                if key_mat == current_mat and key_mat != '':
                    continue
                if merge_start is not None and r - merge_start > 1 and current_mat != '':
                    ws.merge_cells(start_row=merge_start, start_column=prod_col, end_row=r - 1, end_column=prod_col)
                current_mat = key_mat
                merge_start = r

        _merge_same_spec(writer.book['腿肉TSC'], tsc_legs)
        _merge_same_spec(writer.book['胸肉TSC'], tsc_breast)
        _merge_product_group(writer.book['腿肉TSC'], tsc_legs)
        _merge_product_group(writer.book['胸肉TSC'], tsc_breast)

        def _append_semi_block(ws, df_summary):
            if df_summary.empty:
                return
            start_row = ws.max_row + 2
            green_fill = PatternFill(fill_type='solid', fgColor='92D050')
            yellow_fill = PatternFill(fill_type='solid', fgColor='FFFF00')
            header_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')

            # title row
            ws.cell(row=start_row, column=1, value='半成品').fill = green_fill
            ws.cell(row=start_row, column=1).alignment = header_align
            ws.cell(row=start_row, column=4, value='有#N/A号的是没有生产成品')

            left = df_summary.sort_values('物料号').reset_index(drop=True)
            right = df_summary.sort_values('入库数量', ascending=False).reset_index(drop=True)
            n = max(len(left), len(right))
            for i in range(n):
                r = start_row + 1 + i
                if i < len(left):
                    ws.cell(r, 1, left.at[i, '物料号']).alignment = left_align
                    ws.cell(r, 2, left.at[i, '物料描述']).alignment = left_align
                    ws.cell(r, 3, _fmt_int_or_dash(left.at[i, '入库数量'])).alignment = right_align
                    ws.cell(r, 4, left.at[i, '系列']).alignment = left_align
                if i < len(right):
                    ws.cell(r, 5, right.at[i, '物料号']).alignment = left_align
                    ws.cell(r, 6, right.at[i, '物料描述']).alignment = left_align
                    ws.cell(r, 7, _fmt_int_or_dash(right.at[i, '入库数量'])).alignment = right_align
                    ws.cell(r, 8, right.at[i, '系列']).alignment = left_align
                    ws.cell(r, 9, i + 1).alignment = header_align

            total_row = start_row + 1 + n
            ws.cell(row=total_row, column=1, value='半成品合计').fill = yellow_fill
            ws.cell(row=total_row, column=3, value=_fmt_int_or_dash(df_summary['入库数量'].sum())).fill = yellow_fill
            ws.cell(row=total_row, column=7, value=_fmt_int_or_dash(df_summary['入库数量'].sum())).fill = yellow_fill

        def _append_labor_cost_block(ws, bb2_src_df, factor):
            if bb2_src_df is None or bb2_src_df.empty:
                return

            src = bb2_src_df.copy()
            for c in ['入库数量', '实际数量', '实际金额']:
                if c in src.columns:
                    src[c] = pd.to_numeric(src[c], errors='coerce').fillna(0.0)
                else:
                    src[c] = 0.0

            raw_no = src['原料号'].fillna('').astype(str).str.strip() if '原料号' in src.columns else pd.Series('', index=src.index)
            raw_desc = src['原料描述'].fillna('').astype(str).str.strip() if '原料描述' in src.columns else pd.Series('', index=src.index)

            is_header = raw_no.eq('')
            is_labor = raw_no.eq('人工费用') | raw_desc.str.contains('人工', na=False)
            detail = src[~is_header & ~is_labor].copy()
            if detail.empty:
                return

            # 对齐参考口径（与系统成本表人工费用块一致）：
            # 分割前原料: SUMIFS(调整后实际量/额, 原料号=人工费用)
            # 分割后原料: SUMIFS(入库数量, 原料号=人工费用)
            # 碎肉量: -SUMIFS(碎肉量, 原料号=人工费用)
            if '调整后实际量' in src.columns:
                src['调整后实际量'] = pd.to_numeric(src['调整后实际量'], errors='coerce').fillna(0.0)
            if '调整后实际额' in src.columns:
                src['调整后实际额'] = pd.to_numeric(src['调整后实际额'], errors='coerce').fillna(0.0)
            if '碎肉量' in src.columns:
                src['碎肉量'] = pd.to_numeric(src['碎肉量'], errors='coerce').fillna(0.0)

            before_qty = float(src.loc[is_labor, '调整后实际量'].sum())
            before_amt = float(src.loc[is_labor, '调整后实际额'].sum())
            after_qty = float(src.loc[is_labor, '入库数量'].sum())
            scrap_qty = float(-src.loc[is_labor, '碎肉量'].sum())
            if scrap_qty < 0:
                scrap_qty = abs(scrap_qty)

            # 参考表：碎肉单价 = 分割前单价 * 系数(腿肉0.95/胸肉0.7)；碎肉金额=碎肉量*碎肉单价
            before_unit = (before_amt / before_qty) if before_qty else None
            scrap_unit = (before_unit * factor) if before_unit is not None else None
            scrap_amt = (scrap_qty * scrap_unit) if (scrap_qty is not None and scrap_unit is not None) else 0.0
            after_amt = float(before_amt - scrap_amt)
            after_unit = (after_amt / after_qty) if after_qty else None
            util_rate = (after_qty / before_qty) if before_qty else None
            scrap_rate = (scrap_qty / before_qty) if before_qty else None
            water_rate = (1 - util_rate - scrap_rate) if (util_rate is not None and scrap_rate is not None) else None

            start_row = ws.max_row + 2
            green_fill = PatternFill(fill_type='solid', fgColor='92D050')
            red_font = Font(color='FF0000', bold=True)
            header_font = Font(bold=True)
            left_align = Alignment(horizontal='left', vertical='center')
            center_align = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')

            # block header
            for c in range(1, 6):
                ws.cell(row=start_row, column=c).fill = green_fill
            ws.cell(row=start_row, column=1, value='人工费用').font = red_font
            ws.cell(row=start_row, column=1).alignment = left_align
            ws.cell(row=start_row, column=3, value='量').font = header_font
            ws.cell(row=start_row, column=4, value='金额').font = header_font
            ws.cell(row=start_row, column=5, value='单价').font = header_font
            ws.cell(row=start_row, column=3).alignment = center_align
            ws.cell(row=start_row, column=4).alignment = center_align
            ws.cell(row=start_row, column=5).alignment = center_align

            summary_rows = [
                ('分割前原料', before_qty, before_amt, before_unit),
                ('分割后原料', after_qty, after_amt, after_unit),
                ('碎肉', scrap_qty, scrap_amt, scrap_unit),
                ('原料利用率', util_rate, None, None),
                ('碎肉', scrap_rate, None, None),
                ('失水', water_rate, None, None),
            ]

            num_fmt = '#,##0.##;[Red](#,##0.##);-'
            unit_fmt = '0.00;[Red](0.00);-'
            pct_fmt = '0%'
            for i, (name, v1, v2, v3) in enumerate(summary_rows, start=1):
                r = start_row + i
                ws.cell(row=r, column=1, value=name).alignment = left_align
                c3 = ws.cell(row=r, column=3, value=v1 if v1 is not None else None)
                c3.alignment = right_align
                if i <= 3:
                    c3.number_format = num_fmt
                else:
                    c3.number_format = pct_fmt
                    c3.alignment = center_align
                c4 = ws.cell(row=r, column=4, value=v2 if v2 is not None else None)
                c4.alignment = right_align
                c4.number_format = num_fmt
                c5 = ws.cell(row=r, column=5, value=v3 if v3 is not None else None)
                c5.alignment = right_align
                c5.number_format = unit_fmt

            # detail header
            head_row = start_row + 8
            for c in range(1, 10):
                ws.cell(row=head_row, column=c).fill = green_fill
                ws.cell(row=head_row, column=c).font = header_font
                ws.cell(row=head_row, column=c).alignment = center_align
            ws.cell(row=head_row, column=1, value='不含琵琶腿')
            ws.cell(row=head_row, column=2, value='修形前原料及产出碎肉')
            ws.cell(row=head_row, column=3, value='耗用金额')
            ws.cell(row=head_row, column=4, value='耗用量')
            ws.cell(row=head_row, column=5, value='单价')
            ws.cell(row=head_row, column=6, value='调整后耗用量')
            ws.cell(row=head_row, column=7, value='')
            ws.cell(row=head_row, column=8, value='调整后金额')
            ws.cell(row=head_row, column=9, value='')

            detail_grp = (
                detail.groupby(['原料号', '原料描述'], as_index=False)[['实际金额', '实际数量']]
                .sum()
                .sort_values('原料号')
                .reset_index(drop=True)
            )
            total_adj_qty = float(detail_grp.loc[detail_grp['实际数量'] > 0, '实际数量'].sum())
            for i, row in detail_grp.iterrows():
                r = head_row + 1 + i
                qty = float(row['实际数量']) if pd.notna(row['实际数量']) else 0.0
                amt = float(row['实际金额']) if pd.notna(row['实际金额']) else 0.0
                unit = (amt / qty) if qty else None
                adj_qty = qty if qty > 0 else None
                adj_amt = amt if amt > 0 else None
                adj_ratio = (adj_qty / total_adj_qty) if (adj_qty is not None and total_adj_qty > 0) else None
                adj_unit = (adj_amt / adj_qty) if (adj_amt is not None and adj_qty not in (None, 0)) else None
                ws.cell(row=r, column=1, value=row['原料号']).alignment = left_align
                ws.cell(row=r, column=2, value=row['原料描述']).alignment = left_align
                c3 = ws.cell(row=r, column=3, value=amt)
                c3.alignment = right_align
                c3.number_format = num_fmt
                c4 = ws.cell(row=r, column=4, value=qty)
                c4.alignment = right_align
                c4.number_format = num_fmt
                c5 = ws.cell(row=r, column=5, value=unit if unit is not None else None)
                c5.alignment = right_align
                c5.number_format = unit_fmt
                c6 = ws.cell(row=r, column=6, value=adj_qty)
                c6.alignment = right_align
                c6.number_format = num_fmt
                c7 = ws.cell(row=r, column=7, value=adj_ratio)
                c7.alignment = center_align
                c7.number_format = pct_fmt
                c8 = ws.cell(row=r, column=8, value=adj_amt)
                c8.alignment = right_align
                c8.number_format = num_fmt
                c9 = ws.cell(row=r, column=9, value=adj_unit)
                c9.alignment = right_align
                c9.number_format = unit_fmt
                # 与参考表右侧标记列口径一致：正向耗用标记1
                ws.cell(row=r, column=10, value=1 if qty > 0 else None).alignment = center_align

        def _apply_grid_borders(ws):
            thin = Side(border_style='thin', color='000000')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).border = border

        def _coerce_numeric_text_cells(ws):
            # Convert numeric-like text to true numeric values to avoid Excel warnings.
            num_re = re.compile(r'^[+-]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?$')
            paren_re = re.compile(r'^\((?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?\)$')
            pct_re = re.compile(r'^[+-]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?%$')
            pct_paren_re = re.compile(r'^\((?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?%\)$')
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    v = cell.value
                    if not isinstance(v, str):
                        continue
                    s = v.strip()
                    if s in ('', '-', '—'):
                        continue
                    if s.startswith('='):
                        continue

                    is_percent = False
                    neg = False
                    parsed = None

                    if pct_paren_re.fullmatch(s):
                        is_percent = True
                        neg = True
                        s = s[1:-1]  # remove ()
                    elif pct_re.fullmatch(s):
                        is_percent = True
                    elif paren_re.fullmatch(s):
                        neg = True
                        s = s[1:-1]
                    elif not num_re.fullmatch(s):
                        continue

                    if is_percent:
                        s_num = s[:-1]
                    else:
                        s_num = s
                    s_num = s_num.replace(',', '')
                    try:
                        parsed = float(s_num)
                    except Exception:
                        continue
                    if neg:
                        parsed = -parsed
                    if is_percent:
                        parsed = parsed / 100.0
                        # keep minus-sign style for negative percentages (not parentheses)
                        if '.' in s_num:
                            cell.number_format = '0.0%;[Red]-0.0%;-'
                        else:
                            cell.number_format = '0%;[Red]-0%;-'
                    else:
                        if float(parsed).is_integer():
                            cell.number_format = '#,##0;[Red](#,##0);-'
                            parsed = int(parsed)
                        else:
                            cell.number_format = '#,##0.##;[Red](#,##0.##);-'
                    cell.value = parsed

        def _remove_thousand_sep_for_xhyl(ws):
            # TSC: keep "修行后原料" as plain numeric id without thousand separators.
            target_col = None
            for hr in range(1, min(ws.max_row, 6) + 1):
                for c in range(1, ws.max_column + 1):
                    if str(ws.cell(hr, c).value).strip() == '修行后原料':
                        target_col = c
                        break
                if target_col:
                    break
            if not target_col:
                return
            start_row = 5 if ws.title in ('腿肉TSC', '胸肉TSC') else 2
            for r in range(start_row, ws.max_row + 1):
                cell = ws.cell(r, target_col)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = '0'

        def _remove_thousand_sep_for_month(ws):
            # Keep 月份/配方月份 as plain integer id without thousand separators.
            target_cols = []
            for hr in range(1, min(ws.max_row, 6) + 1):
                for c in range(1, ws.max_column + 1):
                    if str(ws.cell(hr, c).value).strip() in {'配方月份', '月份'}:
                        target_cols.append(c)
            if not target_cols:
                return
            start_row = 2
            for c in set(target_cols):
                for r in range(start_row, ws.max_row + 1):
                    cell = ws.cell(r, c)
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.number_format = '0'

        def _center_dash_cells(ws):
            # Ensure dash placeholders are centered consistently.
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(r, c)
                    if isinstance(cell.value, str) and cell.value.strip() == '-':
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        continue
                    # Some cells render "-" via number format zero-section (e.g. #,##0;(#,##0);-).
                    # Keep numeric value unchanged, only force centered alignment for visual consistency.
                    if isinstance(cell.value, (int, float)) and cell.value == 0:
                        nf = str(cell.number_format or '')
                        parts = nf.split(';')
                        if len(parts) >= 3 and '-' in parts[2]:
                            cell.alignment = Alignment(horizontal='center', vertical='center')

        # append block to legs/breast sheets
        if f'{prefix}腿肉' in writer.book.sheetnames:
            _append_labor_cost_block(writer.book[f'{prefix}腿肉'], bb2_legs, 0.95)
            _append_semi_block(writer.book[f'{prefix}腿肉'], _build_semi_summary(bb2_legs))
        if f'{prefix}胸肉' in writer.book.sheetnames:
            _append_labor_cost_block(writer.book[f'{prefix}胸肉'], bb2_breast, 0.70)
            _append_semi_block(writer.book[f'{prefix}胸肉'], _build_semi_summary(bb2_breast))

        # 网格线：TSC + XX腿肉/XX胸肉
        for sheet_name in [f'{prefix}腿肉', f'{prefix}胸肉', '腿肉TSC', '胸肉TSC']:
            if sheet_name in writer.book.sheetnames:
                _apply_grid_borders(writer.book[sheet_name])

        # Final safety pass: convert all numeric-like text to real numbers across workbook.
        for ws in writer.book.worksheets:
            _coerce_numeric_text_cells(ws)
            _remove_thousand_sep_for_xhyl(ws)
            _remove_thousand_sep_for_month(ws)
            _center_dash_cells(ws)

        for ws in writer.book.worksheets:
            if ws.title in ('腿肉TSC', '胸肉TSC'):
                _autofit(ws, min_width=10, max_width=50, padding=4)
            else:
                _autofit(ws)
            # uniform row height
            for r in range(1, ws.max_row + 1):
                ws.row_dimensions[r].height = 18
    return output.getvalue()


if file_compare and file_rawlist and file_q3 and file_map:
    st.session_state.pop('download_data', None)
    try:
        month_label = _month_label_from_filename(file_compare.name)
        quarter_label = _quarter_label_from_filename(file_q3.name)
        month_code = _month_code_from_filename(file_compare.name) or _month_code_from_filename(file_q3.name)
        (
            legs_df,
            breast_df,
            other_df,
            bb2_all,
            bb2_legs,
            bb2_breast,
            raw_usage_legs,
            raw_usage_breast,
            spec_map_legs,
            spec_map_breast,
            finished_df,
            labor_df,
            semi_category,
        ) = compute(
            file_compare,
            file_rawlist,
            file_q3,
            file_map,
            month_label=month_label,
            quarter_label=quarter_label,
            month_code=month_code,
        )
    except Exception as exc:
        st.error(f'计算失败: {exc}')
        # column preview removed per request
    else:
        display_legs = _drop_hidden_cols(legs_df)
        display_breast = _drop_hidden_cols(breast_df)
        display_other = _drop_hidden_cols(other_df)

        st.success('计算完成')
        if not display_legs.empty:
            st.subheader('腿肉')
            st.dataframe(display_legs.style.format(FMT_DISPLAY), use_container_width=True)
        if not display_breast.empty:
            st.subheader('胸肉')
            st.dataframe(display_breast.style.format(FMT_DISPLAY), use_container_width=True)
        if not display_other.empty:
            st.subheader('其他')
            st.dataframe(display_other.style.format(FMT_DISPLAY), use_container_width=True)
        if display_legs.empty and display_breast.empty and display_other.empty:
            st.info('本次计算没有可显示的数据。')

        try:
            prefix = (
                file_compare.name.split('_')[0]
                if '_' in file_compare.name
                else file_compare.name.split('.')[0]
            )
            output_name = _build_output_filename(file_compare.name, file_rawlist.name)
            data = to_excel_bytes(
                display_legs,
                display_breast,
                bb2_all,
                bb2_legs,
                bb2_breast,
                raw_usage_legs,
                raw_usage_breast,
                spec_map_legs,
                spec_map_breast,
                prefix,
                finished_df,
                labor_df,
                month_label,
                semi_category,
            )
            has_any = (
                (not display_legs.empty)
                or (not display_breast.empty)
                or (not display_other.empty)
                or (not bb2_all.empty)
                or (not bb2_legs.empty)
                or (not bb2_breast.empty)
                or (not labor_df.empty)
                or (not raw_usage_legs.empty)
                or (not raw_usage_breast.empty)
            )
            if has_any:
                st.session_state['download_data'] = data
                st.session_state['download_name'] = output_name
        except Exception as exc:
            st.error(f'下载生成失败: {exc}')
else:
    st.info('请先上传四个Excel文件。')

if 'download_data' in st.session_state:
    st.download_button(
        label='下载结果 Excel',
        data=st.session_state['download_data'],
        file_name=st.session_state.get('download_name', '系统成本.xlsx'),
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
